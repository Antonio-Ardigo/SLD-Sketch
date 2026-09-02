#!/usr/bin/env python3
"""Generate the example site-survey workbooks (and a blank template).

Run:  python make_examples.py
Writes examples/config1..6 workbooks and template.xlsx.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ARIAL = "Arial"
HEADERS = ["ID", "Type", "Description", "Rating", "Voltage", "Protection",
           "Feeds From", "Notes"]
COL_WIDTHS = [10, 16, 26, 12, 12, 13, 14, 30]
TYPES = ["MV Incomer", "Generator", "MV Busbar", "RMU", "Transformer",
         "Pump", "LV Busbar", "Feeder", "MCC", "Bus Coupler",
         "Capacitor Bank", "Earthing/NER", "Surge Arrester"]
PROTECTIONS = ["CB", "LBS", "Fuse", "Fuse-switch", "Contactor",
               "Fuse-contactor"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")


def build_workbook(path, info, rows, legend_note=None):
    wb = Workbook()

    # ---- Info sheet ----------------------------------------------------
    ws = wb.active
    ws.title = "Info"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 46
    for r, (key, val) in enumerate(info, start=1):
        kc = ws.cell(row=r, column=1, value=key)
        kc.font = Font(name=ARIAL, bold=True, size=11)
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = Font(name=ARIAL, size=11)
        vc.fill = INPUT_FILL

    # ---- Equipment sheet -----------------------------------------------
    ws = wb.create_sheet("Equipment")
    for j, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(name=ARIAL, bold=True, size=11, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name=ARIAL, size=11)
            c.fill = INPUT_FILL

    # Type column dropdown (rows 2..60 so extra rows added on site get it)
    dv = DataValidation(type="list",
                        formula1='"' + ",".join(TYPES) + '"',
                        allow_blank=True, showErrorMessage=True,
                        errorTitle="Unknown type",
                        error="Pick one of: " + ", ".join(TYPES))
    ws.add_data_validation(dv)
    dv.add("B2:B60")
    # Protection dropdown: suggestions only, free text stays allowed
    # (e.g. "87B differential" on a busbar becomes a label annotation)
    dvp = DataValidation(type="list",
                         formula1='"' + ",".join(PROTECTIONS) + '"',
                         allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dvp)
    dvp.add("F2:F60")
    # extend yellow input fill to the blank rows people will fill on site
    for i in range(len(rows) + 2, 31):
        for j in range(1, len(HEADERS) + 1):
            ws.cell(row=i, column=j).fill = INPUT_FILL
            ws.cell(row=i, column=j).font = Font(name=ARIAL, size=11)

    # ---- legend / how-to sheet -----------------------------------------
    ws = wb.create_sheet("How to fill")
    ws.column_dimensions["A"].width = 100
    lines = [
        "HOW TO FILL THIS WORKBOOK (on site)",
        "",
        "Yellow cells are the ones to fill in. Everything else is fixed.",
        "One row per item, top of the network first. The drawing's topology",
        "comes only from ID and Feeds From: an item is drawn under whatever it",
        "feeds from. Row order sets left-to-right order. A half-filled row still",
        "draws, with an open terminal where the supply or the load is missing.",
        "",
        "COLUMNS",
        "  ID          a short unique tag you invent: MV1, RMU1, TX1, BB1, F1",
        "  Type        pick from the dropdown (see TYPES)",
        "  Description free text, printed as a label (a few words also change",
        "              the symbol - see WORDS)",
        "  Rating      from the nameplate: 1000 kVA, 630 A, 315 kW",
        "  Voltage     from the nameplate: 11/0.4 kV, 400 V - printed only",
        "  Protection  the device on THIS item's supply side: CB, LBS, Fuse,",
        "              Fuse-switch, Contactor, Fuse-contactor; blank = the usual",
        "              default; two supplies: a comma list in Feeds From order",
        "              (LBS, CB). Free text on a busbar (87B differential) is",
        "              printed as a zone label.",
        "  Feeds From  the ID of the item supplying this one; comma for two",
        "              supplies (BB1, BB2 - MV1, MV2)",
        "  Notes       anything worth remembering; printed on ties; a few",
        "              leading words change the symbol - see WORDS",
        "",
        "TYPES (and how each one draws)",
        "  MV Incomer     source tick at the top; feeds an MV Busbar, RMU or TX",
        "  MV Busbar      thick bar, a breaker on every way; a board fed from",
        "                 another board sits one tier lower",
        "  RMU            dashed enclosure, load-break switches on the ways in,",
        "                 fuse-switches on the tee-offs; RMUs that feed from",
        "                 each other form a ring side by side",
        "  Transformer    two circles; step-down or step-up follows Feeds From",
        "  Generator      G circle over the board it feeds, on top of a step-up",
        "                 column, or under a reversed one",
        "  Pump           M 3~ circle: transformer row on MV gear, feeder band",
        "                 on an LV board or an MCC, under a transformer of its own",
        "  LV Busbar      thick bar with its feeders; fed from a Feeder or",
        "                 another LV Busbar it is a sub-board on the row below",
        "  Feeder         drop + device + arrow; on MV gear an outgoing cable",
        "  MCC            labelled box; with Pumps/Feeders under it, a bus of",
        "                 its own on the row below with the motor ways off it",
        "  Bus Coupler    breaker between two busbars of the same kind; between",
        "                 a board and a Generator, a changeover (ATS)",
        "  Capacitor Bank plates to earth (LV board or MV gear); needs no load",
        "  Earthing/NER   resistor box to earth; needs no load",
        "  Surge Arrester arrester box to earth; needs no load",
        "",
        "WIRING RECIPES (Feeds From)",
        "  Step-down: TX feeds from the MV board/RMU; the LV Busbar from TX.",
        "  Step-up from a genset or PV board: TX feeds from the Generator (or",
        "    generation LV Busbar); the MV Busbar/RMU feeds from TX.",
        "  Step-up from a live LV board: TX feeds from that board; the MV",
        "    gear feeds from TX.",
        "  Genset straight onto a board: the board feeds from MV, DG1, DG2;",
        "    the Generator rows have Feeds From blank.",
        "  Standby set on a changeover: a Bus Coupler row ATS feeding from",
        "    MSB3, G1.",
        "  Sub-board (DB): the LV Busbar feeds from the Feeder that supplies",
        "    it, or straight from the main LV Busbar.",
        "  Motors in an MCC: the MCC feeds from the board; each Pump from the MCC.",
        "  Bus tie: a Bus Coupler feeding from BB1, BB2; Notes Normally open.",
        "  Board with two supplies: the board feeds from TX1, TX2.",
        "  Ring of RMUs from a board: each RMU feeds from its neighbours",
        "    (PB, R2 ... R4, PB); a link written on both rows is drawn once.",
        "  Open point of a ring: Notes on one RMU: N.O. towards RMU2.",
        "  Spur or sub-ring off an RMU: the spur RMU feeds from the ring RMU;",
        "    a sub-ring's two RMUs feed from it and from each other.",
        "  Several voltage levels: the lower board feeds from the transformer,",
        "    the transformer from the upper board; tiers follow the wiring,",
        "    never the Voltage column.",
        "  Outgoing MV cable, capacitor bank, NER: a Feeder, Capacitor Bank",
        "    or Earthing/NER row feeding from the MV board or RMU.",
        "",
        "WORDS THAT CHANGE THE SYMBOL (read from Description and Notes)",
        "  VSD / VFD / drive        on a Pump: a drive box on the motor's drop",
        "  Spare / Future /         at the start of Notes: the way drawn dashed",
        "    Out of service",
        "  N.O. / Normally open     on a Bus Coupler: printed under it",
        "  N.O. towards RMU2,       on an RMU: an N.O. marker on the cable to",
        "    ring open here           the RMU named, or under the box",
        "  capacitor / PFC / kvar   on a Feeder: drawn as a capacitor bank",
        "  NER / earthing / zig-zag on a Feeder, or a Transformer with nothing",
        "                           on its output: earthing resistor / tx",
        "  arrester / surge         on a Feeder: drawn as a surge arrester",
        "  Protection words: CB MCCB MCB ACB - LBS isolator - Fuse -",
        "    Fuse-switch SFU - Contactor - Fuse-contactor starter",
        "",
        "BACK AT THE OFFICE",
        "  python sld_sketch.py thisfile.xlsx           -> the SVG",
        "  python sld_sketch.py thisfile.xlsx --dxf     -> SVG + DXF with the",
        "                                                  equipment table",
        "  python sld_check.py thisfile.xlsx            -> checks the drawing",
        "                                                  against the table",
    ]
    if legend_note:
        lines.insert(2, legend_note)
    for r, s in enumerate(lines, start=1):
        c = ws.cell(row=r, column=1, value=s)
        c.font = Font(name=ARIAL, size=11, bold=(r == 1))

    wb.save(path)
    print(f"wrote {path}")


# ------------------------------------------------------------------ data
# row = (ID, Type, Description, Rating, Voltage, Protection, Feeds From, Notes)

CONFIG1 = dict(
    path="examples/config1_single_tx.xlsx",
    info=[("Site", "Example Site A"),
          ("Date", "2026-08-31"),
          ("Surveyed by", "A. Ardigo"),
          ("Notes", "Config 1 - single transformer substation")],
    rows=[
        ("MV1", "MV Incomer", "Utility supply", "", "11 kV", "", "",
         "Cable from utility"),
        ("RMU1", "RMU", "3-way ring main unit", "630 A", "11 kV", "LBS",
         "MV1", ""),
        ("TX1", "Transformer", "Oil-immersed, Dyn11", "1000 kVA",
         "11/0.4 kV", "Fuse-switch", "RMU1", ""),
        ("BB1", "LV Busbar", "Main LV board", "1600 A", "400 V", "CB",
         "TX1", ""),
        ("F1", "Feeder", "Lighting", "100 A", "400 V", "CB", "BB1", ""),
        ("F2", "Feeder", "Small power", "160 A", "400 V", "CB", "BB1", ""),
        ("F3", "Feeder", "HVAC", "250 A", "400 V", "CB", "BB1", ""),
        ("F4", "Feeder", "Spare", "100 A", "400 V", "CB", "BB1", ""),
    ])

CONFIG2 = dict(
    path="examples/config2_twin_tx.xlsx",
    info=[("Site", "Example Site B"),
          ("Date", "2026-08-31"),
          ("Surveyed by", "A. Ardigo"),
          ("Notes", "Config 2 - twin transformers with bus coupler")],
    rows=[
        ("MV1", "MV Incomer", "Utility supply", "", "11 kV", "", "", ""),
        ("RMU1", "RMU", "4-way ring main unit", "630 A", "11 kV", "LBS",
         "MV1", ""),
        ("TX1", "Transformer", "Cast resin, Dyn11", "1600 kVA",
         "11/0.4 kV", "Fuse-switch", "RMU1", ""),
        ("TX2", "Transformer", "Cast resin, Dyn11", "1600 kVA",
         "11/0.4 kV", "Fuse-switch", "RMU1", ""),
        ("BB1", "LV Busbar", "LV board A", "2500 A", "400 V", "CB",
         "TX1", ""),
        ("BB2", "LV Busbar", "LV board B", "2500 A", "400 V", "CB",
         "TX2", ""),
        ("BC1", "Bus Coupler", "Bus section coupler", "2500 A", "400 V",
         "CB", "BB1, BB2", "Normally open"),
        ("F1", "Feeder", "Chillers", "630 A", "400 V", "CB", "BB1", ""),
        ("F2", "Feeder", "Riser 1", "400 A", "400 V", "CB", "BB1", ""),
        ("F3", "Feeder", "Lighting", "160 A", "400 V", "CB", "BB1", ""),
        ("F4", "Feeder", "Pumps", "400 A", "400 V", "CB", "BB2", ""),
        ("F5", "Feeder", "Riser 2", "400 A", "400 V", "CB", "BB2", ""),
        ("F6", "Feeder", "Spare", "250 A", "400 V", "CB", "BB2", ""),
    ])

CONFIG3 = dict(
    path="examples/config3_ring_main.xlsx",
    info=[("Site", "Example Site C"),
          ("Date", "2026-08-31"),
          ("Surveyed by", "A. Ardigo"),
          ("Notes", "Config 3 - ring main, two MV incomers")],
    rows=[
        ("MV1", "MV Incomer", "Ring in", "", "11 kV", "", "",
         "From substation North"),
        ("MV2", "MV Incomer", "Ring out", "", "11 kV", "", "",
         "To substation South"),
        ("RMU1", "RMU", "3-way ring main unit", "630 A", "11 kV",
         "LBS, LBS", "MV1, MV2", ""),
        ("TX1", "Transformer", "Oil-immersed, Dyn11", "800 kVA",
         "11/0.4 kV", "Fuse-switch", "RMU1", ""),
        ("BB1", "LV Busbar", "Main LV board", "1250 A", "400 V", "CB",
         "TX1", ""),
        ("F1", "Feeder", "Distribution board 1", "250 A", "400 V", "CB",
         "BB1", ""),
        ("F2", "Feeder", "Distribution board 2", "250 A", "400 V", "CB",
         "BB1", ""),
        ("F3", "Feeder", "Spare", "160 A", "400 V", "CB", "BB1", ""),
    ])


def _c4_board(mvb, txs, pumps, lvbs):
    rows = []
    for tid, riser, kva in txs:
        rows.append((tid, "Transformer", f"Riser {riser}, Dyn11", kva,
                     "11/0.4 kV", "CB", mvb, ""))
    for pid, desc, kw in pumps:
        rows.append((pid, "Pump", desc, kw, "11 kV", "Fuse-contactor", mvb,
                     ""))
    for bid, riser, amps, tid, mccs in lvbs:
        rows.append((bid, "LV Busbar", f"LV board {riser}", amps, "400 V",
                     "CB", tid, ""))
        for mid, mdesc, mamps in mccs:
            rows.append((mid, "MCC", mdesc, mamps, "400 V", "CB", bid, ""))
    return rows


CONFIG4 = dict(
    path="examples/config4_dual_mv_boards.xlsx",
    info=[("Site", "Example Site D"),
          ("Date", "2026-08-31"),
          ("Surveyed by", "A. Ardigo"),
          ("Notes", "Config 4 - dual MV boards, riser feeders, N.O. tie")],
    rows=[
        ("MV1", "MV Incomer", "Utility incomer A", "", "11 kV", "", "", ""),
        ("MV2", "MV Incomer", "Utility incomer B", "", "11 kV", "", "", ""),
        ("MVB1", "MV Busbar", "MV switchboard A", "630 A", "11 kV", "CB",
         "MV1", "6 feeders to risers"),
        ("MVB2", "MV Busbar", "MV switchboard B", "630 A", "11 kV", "CB",
         "MV2", "6 feeders to risers"),
        ("TIE1", "Bus Coupler", "MV bus tie", "630 A", "11 kV", "CB",
         "MVB1, MVB2", "Normally open"),
    ] + _c4_board("MVB1",
        [("TX1", "R1", "1600 kVA"), ("TX2", "R2", "1250 kVA"),
         ("TX3", "R3", "1000 kVA")],
        [("P1", "CHW pump 1", "315 kW"), ("P2", "CHW pump 2", "315 kW"),
         ("P3", "CW pump 1", "160 kW")],
        [("LVB1", "R1", "2500 A", "TX1",
          [("MCC1", "AHU plant", "400 A"), ("MCC2", "Pump room", "400 A"),
           ("MCC3", "Ventilation", "250 A")]),
         ("LVB2", "R2", "2000 A", "TX2",
          [("MCC4", "AHU plant", "400 A"), ("MCC5", "Ventilation", "250 A")]),
         ("LVB3", "R3", "1600 A", "TX3",
          [("MCC6", "AHU plant", "400 A"), ("MCC7", "Ventilation", "250 A")]),
        ]) + _c4_board("MVB2",
        [("TX4", "R4", "1600 kVA"), ("TX5", "R5", "1250 kVA"),
         ("TX6", "R6", "1000 kVA")],
        [("P4", "CHW pump 3", "315 kW"), ("P5", "CHW pump 4", "315 kW"),
         ("P6", "CW pump 2", "160 kW")],
        [("LVB4", "R4", "2500 A", "TX4",
          [("MCC8", "AHU plant", "400 A"), ("MCC9", "Pump room", "400 A"),
           ("MCC10", "Ventilation", "250 A")]),
         ("LVB5", "R5", "2000 A", "TX5",
          [("MCC11", "AHU plant", "400 A"),
           ("MCC12", "Ventilation", "250 A")]),
         ("LVB6", "R6", "1600 A", "TX6",
          [("MCC13", "AHU plant", "400 A"),
           ("MCC14", "Ventilation", "250 A")]),
        ]))


def _c56_rows(ring_closed):
    rmu3_from = "RMU1, RMU2" if ring_closed else "RMU1"
    rmu3_prot = "LBS, LBS" if ring_closed else "LBS"
    rmu3_note = "Ring closed via RMU2" if ring_closed else ""
    rows = [
        ("MV1", "MV Incomer", "Utility supply", "", "11 kV", "", "", ""),
        ("RMU1", "RMU", "3-way RMU, main", "630 A", "11 kV", "LBS", "MV1",
         "Feeds RMU2 and RMU3"),
        ("RMU2", "RMU", "3-way RMU, substation 2", "630 A", "11 kV", "LBS",
         "RMU1", ""),
        ("RMU3", "RMU", "3-way RMU, substation 3", "630 A", "11 kV",
         rmu3_prot, rmu3_from, rmu3_note),
        ("TX1", "Transformer", "Oil-immersed, Dyn11", "1000 kVA",
         "11/0.4 kV", "Fuse-switch", "RMU2", ""),
        ("TX2", "Transformer", "Oil-immersed, Dyn11", "1000 kVA",
         "11/0.4 kV", "Fuse-switch", "RMU3", ""),
    ]
    for i, (panel, tid) in enumerate(
            [("LVP1", "TX1"), ("LVP2", "TX1"), ("LVP3", "TX2"),
             ("LVP4", "TX2")], start=1):
        rows.append((panel, "LV Busbar", f"LV panel {i}", "800 A", "400 V",
                     "CB", tid, ""))
    for i, panel in enumerate(["LVP1", "LVP2", "LVP3", "LVP4"], start=1):
        rows.append((f"F{2*i-1}", "Feeder", f"Distribution board {i}",
                     "250 A", "400 V", "CB", panel, ""))
        rows.append((f"F{2*i}", "Feeder", "Spare", "160 A", "400 V", "CB",
                     panel, ""))
    return rows


CONFIG5 = dict(
    path="examples/config5_cascaded_rmus.xlsx",
    info=[("Site", "Example Site E"),
          ("Date", "2026-08-31"),
          ("Surveyed by", "A. Ardigo"),
          ("Notes", "Config 5 - three 3-way RMUs, RMU1 feeds RMU2 and RMU3")],
    rows=_c56_rows(ring_closed=False))

CONFIG6 = dict(
    path="examples/config6_closed_ring.xlsx",
    info=[("Site", "Example Site F"),
          ("Date", "2026-08-31"),
          ("Surveyed by", "A. Ardigo"),
          ("Notes", "Config 6 - closed ring: RMU1-RMU2-RMU3-RMU1")],
    rows=_c56_rows(ring_closed=True))

TEMPLATE = dict(
    path="examples/template.xlsx",
    info=[("Site", ""), ("Date", ""), ("Surveyed by", ""), ("Notes", "")],
    rows=[
        ("MV1", "MV Incomer", "Utility supply", "", "11 kV", "", "",
         "EXAMPLE ROW - overwrite with real data"),
    ],
    legend_note="The Equipment sheet contains ONE example row - "
                "overwrite it with the first real item.")


def main():
    os.makedirs("examples", exist_ok=True)
    for cfg in (CONFIG1, CONFIG2, CONFIG3, CONFIG4, CONFIG5, CONFIG6,
                TEMPLATE):
        build_workbook(**cfg)


if __name__ == "__main__":
    main()
