#!/usr/bin/env python3
"""SLD-Sketch — turn a simple site-survey spreadsheet into a single-line diagram.

Usage:
    python sld_sketch.py <workbook.xlsx> [-o output.svg]

The workbook needs two sheets:
  * "Info"       — key/value rows (Site, Date, Surveyed by, Notes)
  * "Equipment"  — one row per item:
        ID | Type | Description | Rating | Voltage | Feeds From | Notes

Recognised types: MV Incomer, RMU, Transformer, LV Busbar, Feeder, Bus Coupler.
"Feeds From" holds the parent item's ID (comma-separated when an item has two
parents, e.g. a bus coupler between two busbars or an RMU on a ring).

Only dependency: openpyxl.
"""

import argparse
import re
import sys
from xml.sax.saxutils import escape

from openpyxl import load_workbook

# ---------------------------------------------------------------- data model

MV_INCOMER = "mv incomer"
RMU = "rmu"
MV_BUSBAR = "mv busbar"
TRANSFORMER = "transformer"
PUMP = "pump"
GENERATOR = "generator"
LV_BUSBAR = "lv busbar"
FEEDER = "feeder"
MCC = "mcc"
BUS_COUPLER = "bus coupler"
CAPACITOR = "capacitor bank"
EARTHING = "earthing"          # neutral earthing resistor / earthing tx
ARRESTER = "surge arrester"
TERMINALS = (CAPACITOR, EARTHING, ARRESTER)   # loads with no outgoing
LV_LOADS = (FEEDER, MCC, PUMP) + TERMINALS    # what hangs off a busbar

TYPE_ALIASES = {
    MV_INCOMER: MV_INCOMER,
    "incomer": MV_INCOMER,
    "mv": MV_INCOMER,
    RMU: RMU,
    "ring main unit": RMU,
    MV_BUSBAR: MV_BUSBAR,
    "mv board": MV_BUSBAR,
    "mv switchboard": MV_BUSBAR,
    "mv distribution board": MV_BUSBAR,
    PUMP: PUMP,
    "motor": PUMP,
    "load": PUMP,
    GENERATOR: GENERATOR,
    "gen": GENERATOR,
    "genset": GENERATOR,
    "gen set": GENERATOR,
    "dg": GENERATOR,
    "alternator": GENERATOR,
    MCC: MCC,
    "motor control centre": MCC,
    "motor control center": MCC,
    TRANSFORMER: TRANSFORMER,
    "trafo": TRANSFORMER,
    "tx": TRANSFORMER,
    "su transformer": TRANSFORMER,     # kept so older sheets still load
    "su tx": TRANSFORMER,
    "su trafo": TRANSFORMER,
    "step-up transformer": TRANSFORMER,
    "step up transformer": TRANSFORMER,
    "step-up": TRANSFORMER,
    "step up": TRANSFORMER,
    LV_BUSBAR: LV_BUSBAR,
    "busbar": LV_BUSBAR,
    "lv board": LV_BUSBAR,
    "lv switchboard": LV_BUSBAR,
    FEEDER: FEEDER,
    "lv feeder": FEEDER,
    "outgoing": FEEDER,
    BUS_COUPLER: BUS_COUPLER,
    "coupler": BUS_COUPLER,
    CAPACITOR: CAPACITOR,
    "capacitor": CAPACITOR,
    "capacitors": CAPACITOR,
    "cap bank": CAPACITOR,
    "pfc": CAPACITOR,
    "power factor correction": CAPACITOR,
    EARTHING: EARTHING,
    "ner": EARTHING,
    "ngr": EARTHING,
    "neutral earthing resistor": EARTHING,
    "neutral grounding resistor": EARTHING,
    "neutral resistor": EARTHING,
    "earthing resistor": EARTHING,
    "earthing transformer": EARTHING,
    "earthing tx": EARTHING,
    "grounding transformer": EARTHING,
    ARRESTER: ARRESTER,
    "arrester": ARRESTER,
    "lightning arrester": ARRESTER,
    "surge": ARRESTER,
}

# words in Description / Notes that say what a row is, whatever its Type
CAP_WORDS = ("capacitor", "pfc", "kvar", "power factor")
EARTH_WORDS = ("ner", "ngr", "earthing", "grounding", "neutral", "zig-zag",
               "zigzag")
ARRESTER_WORDS = ("arrester", "surge")


def words(item):
    return " ".join((item.desc, item.notes)).lower()


def has_word(text, ws):
    return any(re.search(r"(?<![a-z])" + re.escape(w) + r"(?![a-z])", text)
               for w in ws)


def earth_below(items, tx):
    """A transformer with nothing on its output whose row says earthing
    / NER / zig-zag is an earthing transformer: it ends in an earth."""
    return (tx.type == TRANSFORMER
            and not any(tx.id in c.parents for c in items.values())
            and has_word(words(tx), EARTH_WORDS))


def state_words(item):
    """Notes that change how the item's way is drawn, read from the
    start of Notes: spare / future / out of service dash the conductor,
    VSD puts a drive box on a motor's drop, N.O. marks an open way."""
    n = item.notes.strip().lower()
    out = set()
    if n.startswith(("spare", "future", "out of service", "o/s")):
        out.add("spare")
    if has_word(words(item), ("vsd", "vfd", "drive")):
        out.add("vsd")
    if has_word(n, ("n.o.", "n.o", "normally open", "open point",
                    "ring open", "open here")):
        out.add("no")
    return out


PROT_ALIASES = {
    "cb": "cb", "circuit breaker": "cb", "breaker": "cb", "acb": "cb",
    "mccb": "cb", "mcb": "cb", "vcb": "cb",
    "lbs": "lbs", "load break switch": "lbs", "load-break switch": "lbs",
    "switch": "lbs", "isolator": "lbs", "disconnector": "lbs",
    "fuse": "fuse", "fuses": "fuse",
    "fuse-switch": "fuse-switch", "fuse switch": "fuse-switch",
    "switch-fuse": "fuse-switch", "switch fuse": "fuse-switch",
    "sfu": "fuse-switch",
    "contactor": "contactor", "vacuum contactor": "contactor",
    "fuse-contactor": "fuse-contactor", "fuse contactor": "fuse-contactor",
    "fused contactor": "fuse-contactor", "contactor-fuse": "fuse-contactor",
    "motor starter": "fuse-contactor", "starter": "fuse-contactor",
}


def prot_for(item, parent_id=None):
    """The Protection entry for this item's supply from parent_id.

    Returns (raw_text, normalised_kind_or_None). One value applies to
    every supply; a comma list matches the Feeds From order.
    """
    if not item.prots:
        return "", None
    raw = item.prots[0]
    if parent_id is not None and len(item.prots) > 1 \
            and parent_id in item.parents:
        i = item.parents.index(parent_id)
        if i < len(item.prots):
            raw = item.prots[i]
    return raw, PROT_ALIASES.get(" ".join(raw.lower().split()))


class Item:
    def __init__(self, id_, type_, desc, rating, voltage, parents, notes,
                 prots=None):
        self.id = id_
        self.type = type_
        self.desc = desc
        self.rating = rating
        self.voltage = voltage
        self.parents = parents  # list of parent IDs
        self.notes = notes
        self.prots = prots or []  # protection, one entry per supply
        self.x = None  # centre x, assigned during layout
        # busbar geometry
        self.x_left = None
        self.x_right = None
        self.land = {}  # supply id -> x where that supply lands on the bar
        self.tee = {}   # hanging RMU id -> x of its tee-off in this RMU

    def __repr__(self):
        return f"<{self.type} {self.id}>"


def norm_type(raw):
    key = " ".join(str(raw).lower().split())
    return TYPE_ALIASES.get(key)


# ---------------------------------------------------------------- xlsx input

def read_workbook(path):
    wb = load_workbook(path, data_only=True)

    info = {}
    if "Info" in wb.sheetnames:
        for row in wb["Info"].iter_rows(values_only=True):
            if row and row[0]:
                key = str(row[0]).rstrip(":").strip()
                val = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
                info[key.lower()] = val

    if "Equipment" not in wb.sheetnames:
        sys.exit(f"error: no 'Equipment' sheet in {path} "
                 f"(found: {', '.join(wb.sheetnames)})")

    ws = wb["Equipment"]
    rows = list(ws.iter_rows(values_only=True))
    # locate the header row (the one containing "ID" and "Type")
    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c else "" for c in row]
        if "id" in cells and "type" in cells:
            header_idx = i
            headers = cells
            break
    if header_idx is None:
        sys.exit("error: could not find a header row with 'ID' and 'Type' "
                 "columns in the Equipment sheet")

    def col(*names):
        for n in names:
            for j, h in enumerate(headers):
                if n in h:
                    return j
        return None

    c_id = col("id")
    c_type = col("type")
    c_desc = col("desc")
    c_rating = col("rating")
    c_volt = col("volt")
    c_prot = col("protection", "prot")
    c_parent = col("feeds from", "parent", "from")
    c_notes = col("note")

    def cell(row, j):
        if j is None or j >= len(row) or row[j] is None:
            return ""
        return str(row[j]).strip()

    items, order = {}, []
    for row in rows[header_idx + 1:]:
        id_ = cell(row, c_id)
        if not id_:
            continue
        raw_type = cell(row, c_type)
        type_ = norm_type(raw_type)
        if type_ is None:
            print(f"warning: row '{id_}' has unrecognised type '{raw_type}' "
                  f"- drawing it as a generic feeder", file=sys.stderr)
            type_ = FEEDER
        parents = [p.strip() for p in cell(row, c_parent).split(",") if p.strip()]
        prots = [p.strip() for p in cell(row, c_prot).split(",") if p.strip()]
        if id_ in items:
            sys.exit(f"error: duplicate ID '{id_}' in Equipment sheet")
        items[id_] = Item(id_, type_, cell(row, c_desc), cell(row, c_rating),
                          cell(row, c_volt), parents, cell(row, c_notes),
                          prots)
        order.append(id_)

    if not items:
        sys.exit("error: the Equipment sheet has no data rows")

    # a feeder row whose words say capacitor bank / NER / arrester is that
    # item: the symbol follows the words, the Type column can stay Feeder
    for it in items.values():
        if it.type == FEEDER:
            w = words(it)
            if has_word(w, CAP_WORDS):
                it.type = CAPACITOR
            elif has_word(w, ARRESTER_WORDS):
                it.type = ARRESTER
            elif has_word(w, EARTH_WORDS):
                it.type = EARTHING

    for it in items.values():
        for p in it.parents:
            if p not in items:
                sys.exit(f"error: '{it.id}' feeds from unknown ID '{p}' "
                         f"- check the 'Feeds From' column")
        if it.prots and it.type == MV_INCOMER:
            print(f"warning: '{it.id}' - protection on an MV incomer is on "
                  f"the utility side and is not drawn", file=sys.stderr)
        elif it.type not in (LV_BUSBAR, MV_BUSBAR, BUS_COUPLER):
            for raw in it.prots:
                if PROT_ALIASES.get(" ".join(raw.lower().split())) is None:
                    print(f"warning: '{it.id}' - unknown protection "
                          f"'{raw}', the default symbol is drawn",
                          file=sys.stderr)

    for it in items.values():
        if it.type in (TRANSFORMER,):
            up = [c for c in items.values()
                  if it.id in c.parents and c.type in (MV_BUSBAR, RMU)]
            dn = [c for c in items.values()
                  if it.id in c.parents and c.type == LV_BUSBAR]
            if up and dn:
                print(f"warning: '{it.id}' feeds both an MV and an LV board "
                      f"- drawn as a step-up", file=sys.stderr)
        if it.type in (TRANSFORMER,) and not it.parents:
            print(f"warning: '{it.id}' has no Feeds From - drawn with an "
                  f"open supply terminal", file=sys.stderr)
        if it.type == MCC:
            for p in it.parents:
                if p in items and items[p].type in (MV_BUSBAR, RMU):
                    print(f"warning: MCC '{it.id}' feeds from MV gear "
                          f"'{p}' - an MCC is an LV assembly; add a "
                          f"transformer and an LV board in between",
                          file=sys.stderr)
        if it.type == TRANSFORMER and it.parents and not any(
                it.id in c.parents for c in items.values()) \
                and not earth_below(items, it):
            print(f"warning: '{it.id}' has nothing on its output - drawn "
                  f"with an open outgoing terminal; put '{it.id}' in the "
                  f"Feeds From of whatever it supplies", file=sys.stderr)
        if it.type == GENERATOR and not any(it.id in c.parents
                                            for c in items.values()) \
                and not any(items[p].type == TRANSFORMER
                            for p in it.parents if p in items) \
                and not gen_feeds(items, list(items), it):
            print(f"warning: generator '{it.id}' feeds nothing",
                  file=sys.stderr)

    return info, items, order


# ---------------------------------------------------------------- layout

MARGIN = 90
FEEDER_SPACING = 95
BUS_GAP = 110          # horizontal gap between adjacent busbars
MIN_BUS_WIDTH = 170

SLOT_GAP = 30          # gap between slots on an MV switchboard
RMU_TEE = 40           # slot of a tee-off feeding an RMU hung below
PUMP_SLOT = 115        # slot width of a pump/motor way
TX_LABEL_W = 150       # room a transformer's label block needs

Y_LABEL = 34           # MV incomer labels
Y_MV_TOP = 62          # top of the MV incomer stub
Y_RMU_TOP = 150
Y_RMU_BOT = 268
Y_MVBUS = 208          # MV switchboard busbar (same tier as the RMU)
PUMP_R = 20
TX_R = 19
LEGEND_H = 100         # symbol legend strip below the drawing
TIER_H = 150           # vertical pitch of a cascaded MV level
TIER_LINK_H = 200      # pitch when a transformer sits between two tiers
STEPUP_H = 170         # headroom above the MV rows for a step-up column
GEN_H = 110            # headroom for a generator standing over MV gear
STEPUP_SHIFT = 0       # set per drawing by set_tiers()
Y_GEN = 96             # generation source symbol, step-up column
Y_SU_C1 = 196          # step-up transformer, upper circle
Y_SU_C2 = Y_SU_C1 + 27

# rows below the MV distribution - re-based by set_tiers() when MV boards
# or RMUs are fed from another MV board
Y_PUMP = 352           # pump/motor circle centre
Y_TX_C1 = 342          # centre of upper transformer circle
Y_TX_C2 = Y_TX_C1 + 27
Y_BUS = 486
Y_FEED_BRK = 516       # feeder breaker centre
Y_ARROW = 574          # arrow tip
Y_FEED_LBL = 592
DIAG_H = 780
LV_SUB_H = 200         # pitch of an LV sub-board row below its supply board
SUB_PAD = 20           # clearance each side of a sub-board inside its slot


def alloc_lanes(runs, slots):
    """Give every sideways run its own horizontal lane, longest run first
    so nested elbows never cross; runs that do not overlap in x may share a
    lane.  `slots` lists the lane heights in order of preference, chosen so
    no lane passes through the device zone of a drop it crosses; more lanes
    than slots are spread evenly between the first and last slot.
    runs: [(key, x0, x1)].  Returns {key: y}."""
    runs = sorted(runs, key=lambda r: -abs(r[2] - r[1]))
    lanes, idx = [], {}
    for key, x0, x1 in runs:
        lo, hi = min(x0, x1) - 8, max(x0, x1) + 8
        for k, occ in enumerate(lanes):
            if all(hi < a or lo > b for a, b in occ):
                occ.append((lo, hi))
                idx[key] = k
                break
        else:
            lanes.append([(lo, hi)])
            idx[key] = len(lanes) - 1
    n = len(lanes)
    if n <= len(slots):
        ys = list(slots)
    else:
        ys = [slots[0] + (slots[-1] - slots[0]) * k / (n - 1)
              for k in range(n)]
    return {key: ys[k] for key, k in idx.items()}


def set_tiers(extra, top=0, sub_levels=0):
    """Push the transformer/LV rows down by `extra` px; `top` is extra
    headroom above the MV rows for a step-up column; `sub_levels` rows
    of LV sub-boards hang below the LV bar."""
    global Y_PUMP, Y_TX_C1, Y_TX_C2, Y_BUS, Y_FEED_BRK, Y_ARROW
    global Y_FEED_LBL, DIAG_H, STEPUP_SHIFT
    STEPUP_SHIFT = top
    extra = extra + top
    Y_PUMP = 352 + extra
    Y_TX_C1 = 342 + extra
    Y_TX_C2 = Y_TX_C1 + 27
    Y_BUS = 486 + extra
    Y_FEED_BRK = 516 + extra
    Y_ARROW = 574 + extra
    Y_FEED_LBL = 592 + extra
    DIAG_H = 780 + extra + LV_SUB_H * sub_levels


def gen_feeds(items, order, g):
    """The boards a generator supplies, as (board, device, coupler):
    boards naming it in Feeds From, and the far end of a Bus Coupler
    between it and a board (a changeover / ATS)."""
    out = []
    for b in children_of(items, order, g.id, {LV_BUSBAR, MV_BUSBAR, RMU}):
        out.append((b, prot_for(b, g.id)[1] or "cb", None))
    for oid in order:
        c = items[oid]
        if c.type != BUS_COUPLER or g.id not in c.parents:
            continue
        for p in c.parents:
            if p != g.id and items[p].type in (LV_BUSBAR, MV_BUSBAR, RMU):
                out.append((items[p], prot_for(c)[1] or "cb", c))
    return out


def mv_gens(items, order):
    """Generators standing over MV gear as a supply of their own."""
    out = {}
    for oid in order:
        g = items[oid]
        if g.type != GENERATOR:
            continue
        if any(items[p].type == TRANSFORMER for p in g.parents):
            continue                  # the source of a step-up column
        fed = [(b, k, c) for b, k, c in gen_feeds(items, order, g)
               if b.type in (MV_BUSBAR, RMU)]
        if fed:
            out[g.id] = fed
    return out


def su_mid(items, order):
    """Step-ups taking power from a real LV board (one that has its own
    supply) up to an MV board or RMU: drawn in the transformer row, with
    the board below them and their output above."""
    out = {}
    for oid in order:
        tx = items[oid]
        if tx.type not in (TRANSFORMER,):
            continue
        up = children_of(items, order, tx.id, {MV_BUSBAR, RMU})
        src = [items[p] for p in tx.parents
               if items[p].type == LV_BUSBAR and items[p].parents]
        if up and src:
            out[tx.id] = (src[0], up[0])
    return out


def lv_subs(items, order):
    """A transformer taking supply from one LV board
    and feeding another: drawn in the transformer row between the two
    bars, the fed board standing beside its parent with its own feeders."""
    mid = su_mid(items, order)
    out = {}
    for oid in order:
        tx = items[oid]
        if tx.type not in (TRANSFORMER,) or tx.id in mid:
            continue
        src = [items[p] for p in tx.parents
               if items[p].type == LV_BUSBAR and items[p].parents]
        fed = children_of(items, order, tx.id, {LV_BUSBAR})
        if src and fed:
            out[tx.id] = (src[0], fed)
    return out


def step_ups(items, order):
    """Transformers that feed an MV busbar or RMU (step-up), mapped to
    the source row that feeds them."""
    out = {}
    for oid in order:
        tx = items[oid]
        if tx.type not in (TRANSFORMER,):
            continue
        if not children_of(items, order, tx.id, {MV_BUSBAR, RMU}):
            continue
        if any(items[p].type == LV_BUSBAR and items[p].parents
               for p in tx.parents):
            continue                  # drawn in the transformer row instead
        if any(items[p].type in (MV_BUSBAR, RMU) for p in tx.parents):
            continue                  # joins two boards: drawn between tiers
        src = next((items[p] for p in tx.parents), None)
        out[tx.id] = src
    return out


def rmu_hang(items, order):
    """RMUs that hang below another RMU rather than beside it on its
    ring: at an RMU with three or more RMU neighbours, a branch that
    reaches no supply of its own (no incomer, board or transformer) is
    fed only through that RMU, so it is drawn a tier down, as a way of
    it - a spur, or a sub-ring fed at both ends.  {rmu id: {child ids}}"""
    rmus = [i for i in order if items[i].type == RMU]
    adj = {r: set() for r in rmus}
    for r in rmus:
        for p in items[r].parents:
            if items[p].type == RMU:
                adj[r].add(p)
                adj[p].add(r)
    anchored = {r for r in rmus
                if any(items[p].type != RMU for p in items[r].parents)}
    hang = {}
    for v in rmus:
        if len(adj[v]) < 3:
            continue
        for u in adj[v]:
            comp, stack = set(), [u]
            while stack:
                n = stack.pop()
                if n in comp or n == v:
                    continue
                comp.add(n)
                stack.extend(adj[n])
            if not comp & anchored:
                hang.setdefault(v, set()).update(comp & adj[v])
    return hang


def mv_depth(items, order):
    """Vertical tier of each MV busbar / RMU, from Feeds From alone:
    whatever feeds a board is drawn above it.  A board or RMU fed from a
    board, directly or through a transformer, sits one tier below it; RMUs
    linked to each other (a ring) stay level, except a branch that hangs
    off one RMU of a ring (rmu_hang); a board nothing feeds from above is
    a root on the top tier.  Voltage is a label, not a rule."""
    depth = {i: 0 for i in order if items[i].type in (MV_BUSBAR, RMU)}
    hang = rmu_hang(items, order)
    for _ in range(len(depth) + 1):          # relax along the chains
        changed = False
        for oid in depth:
            for p in items[oid].parents:
                par = items[p]
                if par.id in depth:
                    if p in hang.get(oid, ()):
                        continue      # written both ways: p hangs off oid
                    down = par.type == MV_BUSBAR or oid in hang.get(p, ())
                    d = depth[par.id] + (1 if down else 0)
                elif par.type == TRANSFORMER:
                    pp = next((q for q in par.parents if q in depth), None)
                    if pp is None:
                        continue
                    d = depth[pp] + 1
                else:
                    continue
                if d > depth[oid]:
                    depth[oid] = d
                    changed = True
        if not changed:
            break
    return depth


def level_links(items, order, depth=None):
    """Transformers joining two MV boards / RMUs, as (upper, tx, lower)
    with the shallower board first whichever way the row was written."""
    depth = depth if depth is not None else mv_depth(items, order)
    out = []
    for oid in order:
        tx = items[oid]
        if tx.type != TRANSFORMER:
            continue
        ups = [items[p] for p in tx.parents if p in depth]
        downs = children_of(items, order, tx.id, {MV_BUSBAR, RMU})
        for a in ups:
            for b in downs:
                if a is b:
                    continue
                u, l = (a, b) if depth[a.id] <= depth[b.id] else (b, a)
                out.append((u, tx, l))
    return out


def level_tx_ids(items, order):
    return {tx.id for _, tx, _ in level_links(items, order)}


def tier_offsets(depth, links, hung=()):
    """y offset of each tier: 150 px per cascade step, 200 px where a
    transformer sits between the tiers or an RMU hangs off an RMU (its
    feed dog-legs across the gap)."""
    n = max(depth.values(), default=0)
    crossed = {depth[l.id] for _, _, l in links} | {depth[h] for h in hung}
    off, y = [0], 0
    for d in range(1, n + 1):
        y += TIER_LINK_H if d in crossed else TIER_H
        off.append(y)
    return off


def children_of(items, order, pid, types=None):
    out = []
    for oid in order:
        it = items[oid]
        if pid in it.parents and (types is None or it.type in types):
            out.append(it)
    return out


def lv_kids(items, order, bb):
    """The ways of an LV board: feeders, MCCs, motors, terminal items and
    sub-boards fed straight from it, each taking a slot on the bar."""
    return children_of(items, order, bb.id, set(LV_LOADS) | {LV_BUSBAR})


def sub_boards_of(items, order, f):
    """LV boards fed from a feeder: drawn on the row below, under it."""
    return children_of(items, order, f.id, {LV_BUSBAR})


def is_sub_board(items, bb):
    """An LV board fed from a feeder or from another LV board hangs
    below that board and is placed with it, not on the main row."""
    return bb.type == LV_BUSBAR and any(
        items[p].type in (FEEDER, LV_BUSBAR) for p in bb.parents)


def lv_level(items, bb, _seen=None):
    """How many rows below the main LV row a board sits."""
    seen = _seen or set()
    if bb.id in seen:
        return 0
    seen.add(bb.id)
    for p in bb.parents:
        par = items[p]
        if par.type == LV_BUSBAR:
            return lv_level(items, par, seen) + 1
        if par.type == FEEDER:
            for q in par.parents:
                if items[q].type == LV_BUSBAR:
                    return lv_level(items, items[q], seen) + 1
    return 0


def lv_kid_width(items, order, k):
    """Slot width of one way: a plain way, or the sub-board(s) under it."""
    if k.type == LV_BUSBAR:
        return lv_board_width(items, order, k) + 2 * SUB_PAD
    subs = sub_boards_of(items, order, k) if k.type == FEEDER else []
    if subs:
        return max(FEEDER_SPACING,
                   sum(lv_board_width(items, order, b) + 2 * SUB_PAD
                       for b in subs))
    return FEEDER_SPACING


def lv_board_width(items, order, bb):
    kids = lv_kids(items, order, bb)
    return max(MIN_BUS_WIDTH, sum(lv_kid_width(items, order, k)
                                  for k in kids))


def place_lv_board(items, order, bb, center_x):
    """Place an LV busbar on center_x with its ways in their slots, and
    the sub-boards beneath its feeders on the row below."""
    kids = lv_kids(items, order, bb)
    width = lv_board_width(items, order, bb)
    bb.x_left, bb.x_right = center_x - width / 2, center_x + width / 2
    bb.x = center_x
    widths = [lv_kid_width(items, order, k) for k in kids]
    cur = bb.x_left + (width - sum(widths)) / 2
    for k, w in zip(kids, widths):
        k.x = cur + w / 2
        if k.type == LV_BUSBAR:
            place_lv_board(items, order, k, k.x)
        elif k.type == FEEDER:
            subs = sub_boards_of(items, order, k)
            if len(subs) == 1:
                place_lv_board(items, order, subs[0], k.x)
            elif subs:
                c = cur
                for b in subs:
                    bw = lv_board_width(items, order, b) + 2 * SUB_PAD
                    place_lv_board(items, order, b, c + bw / 2)
                    c += bw
        cur += w
    return width


def right_edge(items, x):
    """The cursor to carry on from: past every bar already placed."""
    return max([x] + [b.x_right for b in items.values()
                      if b.type in (LV_BUSBAR, MV_BUSBAR)
                      and b.x_right is not None])


def place_board_row(items, order, boards, x):
    """Place LV boards side by side from x, each sized by its feeders."""
    cur = x + BUS_GAP
    for bb in boards:
        w = lv_board_width(items, order, bb)
        place_lv_board(items, order, bb, cur + w / 2)
        cur += w + BUS_GAP
    return x + BUS_GAP, cur


def place_loose_boards(items, order, x):
    """Transformers and LV boards no other pass claimed - an entry whose
    supply is not filled in yet.  They still get a proper bar with their
    feeders under it, rather than a bare stub in the leftover row."""
    for oid in order:
        tx = items[oid]
        if tx.type not in (TRANSFORMER,) or tx.x is not None:
            continue
        todo = [b for b in children_of(items, order, tx.id, {LV_BUSBAR})
                if b.x is None]
        if not todo:
            continue
        left, cur = place_board_row(items, order, todo,
                                    right_edge(items, x))
        tx.x = (left + cur - BUS_GAP) / 2
        x = cur
    loose = [items[i] for i in order
             if items[i].type == LV_BUSBAR and items[i].x is None
             and not is_sub_board(items, items[i])]
    if loose:
        _, x = place_board_row(items, order, loose, right_edge(items, x))
    for oid in order:                 # a generator feeding an LV board
        g = items[oid]                # stands over that board's bar
        if g.type != GENERATOR or g.x is not None:
            continue
        fed = [b for b, _, _ in gen_feeds(items, order, g)
               if b.type == LV_BUSBAR and b.x_left is not None]
        if not fed:
            continue
        bb = fed[0]
        sup = [items[i] for i in order
               if items[i].type in (TRANSFORMER, GENERATOR)
               and bb.id in [k.id for k in children_of(items, order, i)]
               or items[i].type == GENERATOR and bb.id in
               [b.id for b, _, _ in gen_feeds(items, order, items[i])]]
        n = max(1, len(sup))
        k = sup.index(g) if g in sup else 0
        if is_sub_board(items, bb):   # the feed from above takes the centre
            gk = [s for s in sup if s.type == GENERATOR].index(g)
            g.x = bb.x + 60 * (gk + 1)
        else:
            g.x = (bb.x if n == 1
                   else bb.x_left + (bb.x_right - bb.x_left) * (k + 0.5) / n)
    return x


def place_lv_subs(items, order, x):
    """Place LV/LV transformers and the boards they feed, left to right."""
    for tx_id, (src, fed) in lv_subs(items, order).items():
        todo = [b for b in fed if b.x is None]
        if not todo:
            if items[tx_id].x is None:
                xs = [b.x for b in fed if b.x is not None]
                items[tx_id].x = sum(xs) / len(xs)
            continue
        left, cur = place_board_row(items, order, todo,
                                    right_edge(items, x))
        if items[tx_id].x is None:
            items[tx_id].x = (left + cur - BUS_GAP) / 2
        x = cur
    return x


def slot_width(items, order, item):
    """Width needed under one way of an MV switchboard."""
    if item.type == PUMP:
        return PUMP_SLOT
    if item.type in (TRANSFORMER, GENERATOR):
        boards = children_of(items, order, item.id, {LV_BUSBAR})
        w = 130
        for bb in boards:
            w = max(w, lv_board_width(items, order, bb))
        return w
    if item.type in LV_LOADS:         # an outgoing way of the MV board
        return FEEDER_SPACING
    return 130


def ring_group(items, order, head, depth):
    """RMUs linked to `head` through RMU-to-RMU cables on the same tier,
    in chain order, with the head centred when it has two neighbours."""
    def links(r):
        out = [items[p] for p in r.parents
               if items[p].type == RMU and depth.get(p) == depth.get(r.id)]
        out += [k for k in children_of(items, order, r.id, {RMU})
                if depth.get(k.id) == depth.get(r.id)]
        return out

    seen, chain = {head.id}, []
    for k in links(head):               # walk each branch away from the head
        branch, node = [], k
        while node is not None and node.id not in seen:
            seen.add(node.id)
            branch.append(node)
            node = next((n for n in links(node) if n.id not in seen), None)
        chain.append(branch)
    if len(chain) >= 2:                 # head sits between its two branches
        return list(reversed(chain[0])) + [head] + [n for b in chain[1:]
                                                    for n in b]
    return [head] + [n for b in chain for n in b]


def mv_children(items, order, node):
    """The ways of an MV board / RMU that occupy a slot beneath it.
    An LV-fed step-up is a parent in the graph but a way in the drawing;
    an RMU hanging off this RMU (rmu_hang) is a way too; one member of
    a ring stands for the whole ring, which is laid out as a group."""
    types = {TRANSFORMER, PUMP} | set(LV_LOADS)
    if node.type == MV_BUSBAR:
        types |= {MV_BUSBAR, RMU}
    links = level_links(items, order)
    ltx = {tx.id for _, tx, _ in links}
    kids = [k for k in children_of(items, order, node.id, types)
            if k.id not in ltx]
    if node.type == RMU:
        hang = rmu_hang(items, order).get(node.id, set())
        kids += [k for k in children_of(items, order, node.id, {RMU})
                 if k.id in hang]
    kids += [items[t] for t, (_, up) in su_mid(items, order).items()
             if up.id == node.id and items[t] not in kids]
    seen = set()
    for u, tx, l in links:            # the board on the far side of a
        if u.id == node.id and l not in kids and l.id not in seen:
            kids.append(l)            # transformer hangs beneath this one
            seen.add(l.id)
    depth = mv_depth(items, order)
    out, grouped = [], set()
    for k in kids:
        if k.id in grouped:
            continue
        if k.type == RMU:
            grouped.update(m.id for m in ring_group(items, order, k, depth))
        out.append(k)
    return out


def mv_own_width(items, order, node, depth=None):
    """Width one board / RMU needs for its own ways (no ring members)."""
    if node.type not in (MV_BUSBAR, RMU):
        return slot_width(items, order, node)
    kids = mv_children(items, order, node)
    tees = rmu_hang(items, order).get(node.id, ()) if node.type == RMU \
        else ()
    n_sup = (len(supplies_of(items, order, node, step_ups(items, order)))
             + sum(1 for _, _, l in level_links(items, order)
                   if l.id == node.id))
    top = (n_sup - 1) * 90 + 60 if n_sup > 1 else 0
    if not kids:
        return max(MIN_BUS_WIDTH, top)
    need = (sum(mv_width(items, order, k, depth) for k in kids)
            + SLOT_GAP * (len(kids) - 1)
            + (RMU_TEE + SLOT_GAP) * len(tees))
    return max(MIN_BUS_WIDTH, need, top)


def mv_width(items, order, node, depth=None):
    """Width the whole subtree under an MV board / RMU needs, including
    the ring an RMU belongs to."""
    if node.type == RMU and depth is not None:
        group = ring_group(items, order, node, depth)
        if len(group) > 1:
            return (sum(mv_own_width(items, order, m, depth) for m in group)
                    + SLOT_GAP * (len(group) - 1))
    return mv_own_width(items, order, node, depth)


def place_mv_node(items, order, node, left, width, depth=None):
    """Place an MV board / RMU across [left, left+width] and its ways.
    A ring head lays its whole ring out side by side in that band."""
    if node.type == RMU and depth is not None:
        group = ring_group(items, order, node, depth)
        if len(group) > 1:
            widths = [mv_own_width(items, order, m, depth) for m in group]
            need = sum(widths) + SLOT_GAP * (len(group) - 1)
            cursor = left + (width - need) / 2
            for m, w in zip(group, widths):
                place_own(items, order, m, cursor, w, depth)
                cursor += w + SLOT_GAP
            return
    place_own(items, order, node, left, width, depth)


def place_own(items, order, node, left, width, depth=None):
    """Place one board / RMU and the ways directly beneath it.  An RMU
    hung below an RMU gets a narrow tee-off slot in the enclosure and
    its own subtree beside the enclosure's ways; the enclosure centres
    on its ways, not on the whole band."""
    node.x = left + width / 2
    if node.type == MV_BUSBAR:
        node.x_left, node.x_right = left, left + width
    kids = mv_children(items, order, node)
    tees = ([k for k in children_of(items, order, node.id, {RMU})
             if k.id in rmu_hang(items, order).get(node.id, ())]
            if node.type == RMU else [])
    own = [k for k in kids if k.type != RMU or not tees]
    groups = [k for k in kids if k not in own]
    slots = [(k, mv_width(items, order, k, depth), "way") for k in own] \
        + [(t, RMU_TEE, "tee") for t in tees] \
        + [(g, mv_width(items, order, g, depth), "group") for g in groups]
    need = sum(w for _, w, _ in slots) + SLOT_GAP * max(0, len(slots) - 1)
    cursor = left + (width - need) / 2
    span = []
    for k, w, what in slots:
        if what == "tee":
            node.tee[k.id] = cursor + w / 2
            span.append(cursor + w / 2)
        elif k.type in (MV_BUSBAR, RMU):
            place_mv_node(items, order, k, cursor, w, depth)
        else:
            k.x = cursor + w / 2
            span.append(k.x)
        cursor += w + SLOT_GAP
    if tees and span:
        node.x = (min(span) + max(span)) / 2


def supplies_of(items, order, board, sus):
    """Everything standing over a board as a supply of its own: utility
    incomers, step-up columns and generators feeding it directly."""
    gens = mv_gens(items, order)
    return [items[i] for i in order
            if ((items[i].type == MV_INCOMER or i in sus)
                and board.id in [k.id for k in children_of(items, order, i)])
            or (i in gens and board.id in [b.id for b, _, _ in gens[i]])
            or (items[i].type == MV_BUSBAR and i in board.parents)]


def spread_supplies(items, order, boards, sus, links):
    """Everything that lands on a board from above shares one spread
    over its centre, 90 px apart: transformers from a higher tier,
    utility incomers, step-up columns, generators, and the drop from a
    board above (which keeps its own x and only moves its landing).
    The first board an item feeds decides where it stands."""
    done = set()
    for mvb in boards:
        if mvb.x is None:
            continue
        feeds = [tx for u, tx, l in links if l.id == mvb.id] \
            + supplies_of(items, order, mvb, sus)
        n = len(feeds)
        for i, f in enumerate(feeds):
            x = mvb.x + (i - (n - 1) / 2) * 90
            if f.type == MV_BUSBAR:
                mvb.land[f.id] = x
            elif f.id not in done:
                f.x = x
                done.add(f.id)
                src = sus.get(f.id)
                if src is not None:         # a step-up's source follows it
                    src.x = x
                    if src.type == LV_BUSBAR:
                        src.x_left, src.x_right = x - 85, x + 85


def place_tx_motors(items, order):
    """A motor fed straight from a transformer hangs under it."""
    for oid in order:
        t = items[oid]
        if t.type == TRANSFORMER and t.x is not None:
            for m in children_of(items, order, t.id, {PUMP}):
                m.x = t.x


def place_su_mid(items, order, x):
    """Place an LV-fed step-up and the MV gear above it, left to right."""
    for tx_id, (src, up) in su_mid(items, order).items():
        tx = items[tx_id]
        if tx.x is not None:
            continue
        if up.x is not None:          # the gear already has its slot
            tx.x = up.x
            continue
        tx.x = x + 60
        up.x = tx.x
        if up.type == MV_BUSBAR:
            up.x_left, up.x_right = tx.x - 85, tx.x + 85
        x += 200
    return x


def place_step_ups(items, order):
    """A step-up transformer (and its source) or a generator sits above
    the MV busbar or RMU it feeds, beside any utility incomers there."""
    sus = step_ups(items, order)
    gens = mv_gens(items, order)
    cols = list(sus.items()) + [(g, None) for g in gens]
    for tx_id, src in cols:
        tx = items[tx_id]
        if tx_id in gens:
            fed = [b for b, _, _ in gens[tx_id]]
        else:
            fed = children_of(items, order, tx_id, {MV_BUSBAR, RMU})
        anchor = next((f for f in fed if f.x is not None), None)
        if anchor is None:
            continue
        peers = supplies_of(items, order, anchor, sus)
        n = max(1, len(peers))
        k = peers.index(tx) if tx in peers else 0
        tx.x = anchor.x + (k - (n - 1) / 2) * 90
        if src is not None:
            src.x = tx.x
            if src.type == LV_BUSBAR:
                src.x_left, src.x_right = tx.x - 85, tx.x + 85


def gen_below(items, order, tx):
    """A transformer hung under MV gear whose load is a Generator can only
    be a step-up drawn upside down: a generator is never a load."""
    return (tx.type == TRANSFORMER
            and any(items[p].type in (MV_BUSBAR, RMU) for p in tx.parents)
            and bool(children_of(items, order, tx.id, {GENERATOR})))


def place_su_sources(items, order):
    """The generation source drawn under a reversed step-up follows it."""
    for oid in order:
        tx = items[oid]
        if tx.x is None or not gen_below(items, order, tx):
            continue
        for src in children_of(items, order,
                               tx.id, {GENERATOR, LV_BUSBAR, MV_INCOMER}):
            src.x = tx.x
            if src.type == LV_BUSBAR:   # size the bar from its feeders
                place_lv_board(items, order, src, tx.x)
            else:
                src.x = tx.x


def layout_mv_boards(items, order):
    """Layout when the site has MV switchboards (MV Busbar rows)."""
    mvbs = [items[i] for i in order if items[i].type == MV_BUSBAR]
    mvs = [items[i] for i in order if items[i].type == MV_INCOMER]

    # roots are the boards not fed from another MV board
    depth = mv_depth(items, order)
    links = level_links(items, order, depth)
    lowers = {l.id for _, _, l in links}
    roots = [b for b in mvbs
             if not any(items[p].type == MV_BUSBAR for p in b.parents)
             and b.id not in lowers]
    # an RMU tree of its own (fed by an incomer, not by a board) is a root
    # too, so its ways and any lower-tier board beneath it get slots
    rmus = [items[i] for i in order if items[i].type == RMU]
    roots += [r for r in rmus
              if not any(items[p].type in (MV_BUSBAR, RMU) for p in r.parents)
              and r.id not in lowers and r.x is None]
    x = MARGIN
    for mvb in roots:
        w = mv_width(items, order, mvb, depth)
        place_mv_node(items, order, mvb, x, w, depth)
        x = (mvb.x_right if mvb.x_right is not None else mvb.x + w / 2) \
            + BUS_GAP
    seen = {}
    for u, tx, l in links:            # a level transformer stands over
        if l.x is None:               # the board it joins from above
            continue
        k = seen.get(l.id, 0)
        tx.x = l.x + 70 * k
        seen[l.id] = k + 1

    # LV boards centred under their supply transformer(s) - the mean of
    # the supplies when a board has two incomers
    for oid in order:
        bb = items[oid]
        if bb.type != LV_BUSBAR or bb.x is not None \
                or is_sub_board(items, bb):
            continue
        pxs = [items[p].x for p in bb.parents
               if items[p].type in (TRANSFORMER,)
               and items[p].x is not None]
        if pxs:
            place_lv_board(items, order, bb, sum(pxs) / len(pxs))

    x = place_lv_subs(items, order, x)

    # step-up chains take an incomer position over the board they feed
    place_step_ups(items, order)
    place_su_sources(items, order)
    place_tx_motors(items, order)

    # incomers, step-up columns, generators and feeds from above share
    # one spread over each board
    sus = step_ups(items, order)
    spread_supplies(items, order, mvbs + rmus, sus, links)

    x = place_su_mid(items, order, x)

    x = place_loose_boards(items, order, x)

    # anything left over (e.g. an RMU branch mixed in) goes after the boards
    for oid in order:
        it = items[oid]
        if it.x is None:
            it.x = x + 40
            if it.type in (LV_BUSBAR, MV_BUSBAR):
                it.x_left, it.x_right = it.x - 85, it.x + 85
            x += 130

    return max(x - BUS_GAP + MARGIN, 640) + 230


def sub_levels(items, order):
    return max([lv_level(items, items[i]) for i in order
                if items[i].type == LV_BUSBAR] + [0])


def layout(items, order):
    """Assign x coordinates to every item. Returns total drawing width."""
    if any(items[i].type == MV_BUSBAR for i in order) \
            or max(mv_depth(items, order).values(), default=0) > 0:
        return layout_mv_boards(items, order)
    busbars = [items[i] for i in order if items[i].type == LV_BUSBAR
               and not is_sub_board(items, items[i])]
    rmus = [items[i] for i in order if items[i].type == RMU]
    txs = [items[i] for i in order if items[i].type == TRANSFORMER]
    mvs = [items[i] for i in order if items[i].type == MV_INCOMER]

    # 1. busbars left-to-right, width driven by feeder count
    x = MARGIN
    for bb in busbars:
        width = lv_board_width(items, order, bb)
        place_lv_board(items, order, bb, x + width / 2)
        x = bb.x_right + BUS_GAP

    # a transformer feeding only a motor gets its own slot beside the boards
    for tx in txs:
        if tx.x is None and not children_of(items, order, tx.id, {LV_BUSBAR}) \
                and children_of(items, order, tx.id, {PUMP}):
            tx.x = x + PUMP_SLOT / 2
            x += PUMP_SLOT + BUS_GAP

    # 2. transformers centred over the busbar they feed
    for tx in txs:
        fed = children_of(items, order, tx.id, {LV_BUSBAR})
        if fed:
            siblings = [t for t in txs
                        if children_of(items, order, t.id, {LV_BUSBAR}) == fed]
            if len(siblings) > 1:  # several transformers on one busbar
                k = siblings.index(tx)
                spread = 90
                tx.x = fed[0].x + (k - (len(siblings) - 1) / 2) * spread
            else:  # centred over its panel(s) - a TX may feed two boards
                tx.x = sum(bb.x for bb in fed) / len(fed)

    # 3. RMUs centred over their transformer children
    place_tx_motors(items, order)

    for rmu in rmus:
        # pump ways sit beside the transformer ways, left to right
        pumps_r = children_of(items, order, rmu.id, {PUMP})
        txs_r = children_of(items, order, rmu.id, {TRANSFORMER})
        placed = [k.x for k in txs_r if k.x is not None]
        if pumps_r and placed:
            # clear the last transformer's label block before the first pump
            cursor = max(placed) + TX_R + TX_LABEL_W + PUMP_SLOT / 2
            for k in pumps_r:
                k.x = cursor
                cursor += PUMP_SLOT + SLOT_GAP
        kids = [k for k in txs_r + pumps_r if k.x is not None]
        if kids:
            rmu.x = sum(k.x for k in kids) / len(kids)
    for rmu in rmus:  # cascaded RMU: centre over the RMUs it feeds
        if rmu.x is None:
            kids = children_of(items, order, rmu.id, {RMU})
            placed = [k.x for k in kids if k.x is not None]
            if placed:
                rmu.x = sum(placed) / len(placed)

    # 4. MV incomers spread over the RMU (or over the transformer they feed)
    for rmu in rmus:
        if rmu.x is None:
            continue
        feeds = [m for m in mvs if rmu.id in
                 [c.id for c in children_of(items, order, m.id)]]
        n = len(feeds)
        for i, m in enumerate(feeds):
            m.x = rmu.x + (i - (n - 1) / 2) * 80
    for m in mvs:
        if m.x is None:
            kids = children_of(items, order, m.id)
            placed = [k.x for k in kids if k.x is not None]
            if placed:
                m.x = sum(placed) / len(placed)

    x = place_su_mid(items, order, x)
    x = place_lv_subs(items, order, x)
    x = place_loose_boards(items, order, x)

    # 5. anything still unplaced goes in a row after the busbars
    for oid in order:
        it = items[oid]
        if it.x is None:
            it.x = x + 40
            x += 130

    width = max(x - BUS_GAP + MARGIN, 640)
    width += 230  # room for side labels + title block
    return width


# ---------------------------------------------------------------- SVG symbols

class SVG:
    """The drawing surface: every symbol is built from these primitives,
    so another back-end (the DXF writer) only overrides them.  `layer`
    is a hint the render sets while drawing the frame and the legend."""

    def __init__(self):
        self.parts = []
        self.layer = "drawing"

    def document(self, width, height):
        body = "\n".join(self.parts)
        return (f'<svg xmlns="http://www.w3.org/2000/svg" width="{width:.0f}" '
                f'height="{height}" viewBox="0 0 {width:.0f} {height}">\n'
                f'<rect width="100%" height="100%" fill="white"/>\n'
                f'{body}\n</svg>\n')

    def line(self, x1, y1, x2, y2, w=2, dash=None):
        d = f' stroke-dasharray="{dash}"' if dash else ""
        self.parts.append(
            f'<line x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" y2="{y2:.1f}" '
            f'stroke="#111" stroke-width="{w}"{d} stroke-linecap="round"/>')

    def rect(self, x, y, w, h, sw=2, dash=None, fill="none"):
        d = f' stroke-dasharray="{dash}"' if dash else ""
        self.parts.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{w:.1f}" height="{h:.1f}" '
            f'fill="{fill}" stroke="#111" stroke-width="{sw}"{d}/>')

    def circle(self, x, y, r, sw=2):
        self.parts.append(
            f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{r}" fill="none" '
            f'stroke="#111" stroke-width="{sw}"/>')

    def dot(self, x, y, r=3.2):
        self.parts.append(
            f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{r}" fill="#111"/>')

    def poly(self, pts, fill="#111"):
        p = " ".join(f"{a:.1f},{b:.1f}" for a, b in pts)
        self.parts.append(f'<polygon points="{p}" fill="{fill}"/>')

    def text(self, x, y, s, size=12, anchor="middle", bold=False,
             rotate=None, color="#111"):
        if not s:
            return
        wgt = ' font-weight="bold"' if bold else ""
        tr = (f' transform="translate({x:.1f},{y:.1f}) rotate({rotate})"'
              if rotate is not None else "")
        xy = 'x="0" y="0"' if rotate is not None else f'x="{x:.1f}" y="{y:.1f}"'
        self.parts.append(
            f'<text {xy}{tr} font-family="Arial, Helvetica, sans-serif" '
            f'font-size="{size}" fill="{color}" text-anchor="{anchor}"{wgt}>'
            f'{escape(str(s))}</text>')

    def path(self, d, sw=2):
        self.parts.append(
            f'<path d="{d}" fill="none" stroke="#111" stroke-width="{sw}" '
            f'stroke-linecap="round"/>')

    # -- composite symbols ------------------------------------------------

    def load_break_switch(self, x, ytop, ybot):
        """IEC switch-disconnector: hinge circle, blade, contact bar."""
        self.line(x, ytop, x, ytop + 3)
        self.circle(x, ytop + 5.5, 2.5, sw=1.5)          # hinge = switch
        self.line(x + 2, ytop + 8, x + 9, ybot - 9)      # blade
        self.line(x - 6, ybot - 9, x + 6, ybot - 9)      # contact bar
        self.line(x, ybot - 9, x, ybot)

    def fuse_switch(self, x, ytop, ybot):
        """Switch with a fuse rectangle on the lower run (RMU tee-off)."""
        mid = (ytop + ybot) / 2
        self.load_break_switch(x, ytop, mid + 4)
        self.rect(x - 4, mid + 6, 8, ybot - mid - 8)
        self.line(x, mid + 4, x, mid + 6)
        self.line(x, ybot - 2, x, ybot)

    def device(self, kind, x, y):
        """Protection device centred at y on a vertical conductor.
        Returns the half-height the conductor must leave clear."""
        if kind == "fuse":
            self.rect(x - 4, y - 11, 8, 22)
            self.line(x, y - 11, x, y + 11)
            return 11
        if kind == "lbs":
            self.load_break_switch(x, y - 13, y + 13)
            return 13
        if kind == "fuse-switch":
            self.fuse_switch(x, y - 16, y + 16)
            return 16
        if kind == "contactor":
            # IEC: switch with the arc function symbol at the hinge,
            # open side facing the blade
            self.line(x, y - 13, x, y - 11)
            self.path(f"M {x-4:.1f},{y-7:.1f} A 4,4 0 0 1 "
                      f"{x+4:.1f},{y-7:.1f}")                    # hinge arc
            self.line(x + 2, y - 5, x + 8, y + 7)                # blade
            self.line(x, y + 9, x, y + 13)
            return 13
        if kind == "fuse-contactor":
            # MV motor starter: back-up fuse in series with the contactor
            self.rect(x - 4, y - 16, 8, 12)                      # fuse
            self.line(x, y - 16, x, y - 4)
            self.path(f"M {x-4:.1f},{y:.1f} A 4,4 0 0 1 "
                      f"{x+4:.1f},{y:.1f}")                      # hinge arc
            self.line(x + 2, y + 2, x + 7, y + 11)               # blade
            self.line(x, y + 12, x, y + 16)
            return 16
        # 'cb' and anything unknown - circuit breaker: the switch
        # symbol with an X at its hinge edge (IEC 60617)
        self.line(x, y - 13, x, y - 11)
        self.line(x - 3.5, y - 11, x + 3.5, y - 4)               # X
        self.line(x - 3.5, y - 4, x + 3.5, y - 11)
        self.line(x + 2, y - 4.5, x + 9, y + 4)                  # blade
        self.line(x - 6, y + 4, x + 6, y + 4)                    # contact bar
        self.line(x, y + 4, x, y + 13)
        return 13

    def device_h(self, kind, x, y):
        """Protection device centred at x on a horizontal conductor.
        Returns the half-width the conductor must leave clear."""
        if kind == "fuse":
            self.rect(x - 11, y - 4, 22, 8)
            self.line(x - 11, y, x + 11, y)
            return 11
        if kind == "fuse-switch":
            self.line(x - 16, y, x - 14.5, y)
            self.circle(x - 12, y, 2.5, sw=1.5)     # hinge = switch
            self.line(x - 10, y - 1.5, x + 1, y - 9)  # blade
            self.line(x + 1, y - 6, x + 1, y + 6)   # contact bar
            self.rect(x + 3, y - 4, 12, 8)          # fuse
            self.line(x + 1, y, x + 3, y)
            self.line(x + 15, y, x + 16, y)
            return 16
        if kind == "contactor":
            # IEC: switch with the arc function symbol at the hinge
            self.line(x - 13, y, x - 11, y)
            self.path(f"M {x-7:.1f},{y-4:.1f} A 4,4 0 0 0 "
                      f"{x-7:.1f},{y+4:.1f}")                    # hinge arc
            self.line(x - 5, y - 2, x + 7, y - 8)                # blade
            self.line(x + 9, y, x + 13, y)
            return 13
        if kind == "fuse-contactor":
            self.rect(x - 16, y - 4, 12, 8)                      # fuse
            self.line(x - 16, y, x - 4, y)
            self.path(f"M {x:.1f},{y-4:.1f} A 4,4 0 0 0 "
                      f"{x:.1f},{y+4:.1f}")                      # hinge arc
            self.line(x + 2, y - 2, x + 11, y - 7)               # blade
            self.line(x + 12, y, x + 16, y)
            return 16
        if kind == "lbs":
            self.line(x - 13, y, x - 10.5, y)
            self.circle(x - 8, y, 2.5, sw=1.5)      # hinge = switch
            self.line(x - 6, y - 1.5, x + 4, y - 9)  # blade
            self.line(x + 4, y - 6, x + 4, y + 6)   # contact bar
            self.line(x + 4, y, x + 13, y)
            return 13
        # 'cb' and anything unknown - circuit breaker: the switch
        # symbol with an X at its hinge edge (IEC 60617)
        self.line(x - 13, y, x - 11, y)
        self.line(x - 11, y - 3.5, x - 4, y + 3.5)               # X
        self.line(x - 11, y + 3.5, x - 4, y - 3.5)
        self.line(x - 4.5, y - 2, x + 4, y - 9)                  # blade
        self.line(x + 4, y - 6, x + 4, y + 6)                    # contact bar
        self.line(x + 4, y, x + 13, y)
        return 13

    def drop(self, x, ytop, ybot, kind, ydev=None, dash=None):
        """Vertical conductor with a protection device on it."""
        y = ydev if ydev is not None else (ytop + ybot) / 2
        gap = self.device(kind, x, y)
        self.line(x, ytop, x, y - gap, dash=dash)
        self.line(x, y + gap, x, ybot, dash=dash)

    def transformer(self, x, label_lines, side="right", y=None):
        c1 = Y_TX_C1 if y is None else y
        self.circle(x, c1, TX_R, sw=2.2)
        self.circle(x, c1 + 27, TX_R, sw=2.2)
        ty = c1 - 6
        for s in label_lines:
            if side == "left":
                self.text(x - TX_R - 10, ty, s, anchor="end")
            else:
                self.text(x + TX_R + 10, ty, s, anchor="start")
            ty += 15

    def open_end(self, x, y_from, y_to, note):
        """An unterminated conductor: a stub to an open terminal bar."""
        self.line(x, y_from, x, y_to)
        self.line(x - 12, y_to, x + 12, y_to)
        self.text(x, y_to - 8 if y_to < y_from else y_to + 18, note, size=9)

    def arrow_down(self, x, ytip):
        self.poly([(x - 6, ytip - 11), (x + 6, ytip - 11), (x, ytip)])

    def earth(self, x, y):
        """Earth: three shortening bars under a conductor ending at y."""
        self.line(x - 9, y, x + 9, y)
        self.line(x - 6, y + 4, x + 6, y + 4)
        self.line(x - 3, y + 8, x + 3, y + 8)

    def capacitor(self, x, y):
        """Capacitor bank to earth, plates from y down. Height 24."""
        self.line(x - 9, y, x + 9, y, w=2.5)
        self.line(x - 9, y + 6, x + 9, y + 6, w=2.5)
        self.line(x, y + 6, x, y + 16)
        self.earth(x, y + 16)
        return 24

    def resistor(self, x, y):
        """Neutral earthing resistor to earth, box from y down. Height 46."""
        self.rect(x - 6, y, 12, 30)
        self.line(x, y + 30, x, y + 38)
        self.earth(x, y + 38)
        return 46

    def arrester(self, x, y):
        """Surge arrester to earth, box from y down. Height 44."""
        self.rect(x - 7, y, 14, 28)
        self.line(x, y + 4, x, y + 20)                 # the arrow inside
        self.line(x - 4, y + 16, x, y + 22)
        self.line(x + 4, y + 16, x, y + 22)
        self.line(x, y + 28, x, y + 36)
        self.earth(x, y + 36)
        return 44

    def terminal(self, kind, x, y):
        """A load with no outgoing, hung from a conductor ending at y."""
        if kind == CAPACITOR:
            return self.capacitor(x, y)
        if kind == ARRESTER:
            return self.arrester(x, y)
        return self.resistor(x, y)

    def vsd(self, x, y):
        """Drive box over a motor's drop, the conductor running through."""
        self.rect(x - 12, y - 7, 24, 14, sw=1.5, fill="white")
        self.text(x, y + 3, "VSD", size=7.5)


# ---------------------------------------------------------------- rendering

LEGEND_ITEMS = [
    ("cb", "Circuit breaker"), ("lbs", "Load-break switch"),
    ("fuse", "Fuse"), ("fuse-switch", "Fuse-switch"),
    ("contactor", "Contactor"), ("fuse-contactor", "Fused contactor"),
    ("tx", "Transformer"), ("gen", "Generator"), ("pump", "Pump/motor"), ("bus", "Busbar"),
    ("mcc", "MCC"), ("feeder", "Feeder"), ("rmu", "RMU enclosure"),
]


EXTRA_LEGEND = [(CAPACITOR, "Capacitor bank"), (EARTHING, "Earthing/NER"),
                (ARRESTER, "Surge arrester")]


def draw_legend(svg, extra=()):
    cell = 68
    x0, y0 = 24, DIAG_H + 6
    entries = LEGEND_ITEMS + [e for e in EXTRA_LEGEND if e[0] in extra]
    svg.rect(x0, y0, 16 + cell * len(entries), 82, sw=1.2)
    svg.text(x0 + 8, y0 + 14, "LEGEND", size=10, anchor="start", bold=True)
    ytop, ybot = y0 + 22, y0 + 52
    yc = (ytop + ybot) / 2
    for i, (kind, label) in enumerate(entries):
        cx = x0 + 8 + cell * i + cell / 2
        if kind in ("cb", "fuse", "contactor", "fuse-contactor"):
            svg.drop(cx, ytop, ybot, kind)
        elif kind == "lbs":
            svg.load_break_switch(cx, ytop, ybot)
        elif kind == "fuse-switch":
            svg.fuse_switch(cx, ytop, ybot)
        elif kind == "tx":
            svg.circle(cx, yc - 5, 8)
            svg.circle(cx, yc + 5, 8)
        elif kind == "gen":
            svg.circle(cx, yc, 9)
            svg.text(cx, yc + 3.5, "G", size=9, bold=True)
        elif kind == "pump":
            svg.circle(cx, yc, 9)
            svg.text(cx, yc + 3.5, "M", size=9, bold=True)
        elif kind == "bus":
            svg.line(cx - 14, yc, cx + 14, yc, w=5)
        elif kind == "mcc":
            svg.rect(cx - 11, yc - 8, 22, 16, sw=1.5)
            svg.text(cx, yc + 3, "MCC", size=6.5)
        elif kind == "feeder":
            svg.line(cx, ytop + 2, cx, ybot - 11)
            svg.arrow_down(cx, ybot)
        elif kind == "rmu":
            svg.rect(cx - 13, yc - 10, 26, 20, sw=1.2, dash="4 3")
        elif kind == CAPACITOR:
            svg.line(cx, ytop, cx, ytop + 4)
            svg.capacitor(cx, ytop + 4)
        elif kind == EARTHING:
            svg.rect(cx - 5, ytop, 10, 18)
            svg.line(cx, ytop + 18, cx, ytop + 22)
            svg.earth(cx, ytop + 22)
        elif kind == ARRESTER:
            svg.rect(cx - 6, ytop, 12, 18)
            svg.line(cx, ytop + 3, cx, ytop + 14)
            svg.line(cx - 3, ytop + 11, cx, ytop + 15)
            svg.line(cx + 3, ytop + 11, cx, ytop + 15)
            svg.line(cx, ytop + 18, cx, ytop + 22)
            svg.earth(cx, ytop + 22)
        ty = y0 + 64
        for s in (label.split(" ", 1) if len(label) > 11 else [label]):
            svg.text(cx, ty, s, size=9)
            ty += 10


def render(info, items, order, width, canvas=None):
    """Draw the sheet onto `canvas` (an SVG by default, or any object with
    the same primitives) and return its document."""
    svg = canvas if canvas is not None else SVG()
    depth = mv_depth(items, order)
    sus = step_ups(items, order)
    gens = mv_gens(items, order)
    hang = rmu_hang(items, order)
    links = level_links(items, order, depth)
    ltx = {tx.id for _, tx, _ in links}
    tier_off = tier_offsets(depth, links,
                            [c for kids in hang.values() for c in kids]
                            + [b.id for fed in gens.values()
                               for b, _, _ in fed if b.id in depth])
    set_tiers(tier_off[-1] if tier_off else 0,
              STEPUP_H if sus else GEN_H if gens else 0,
              sub_levels(items, order))

    def dy(it):                       # vertical offset of an MV tier
        return tier_off[depth.get(it.id, 0)] if tier_off else 0

    def lv_y(bb):                     # bar level of an LV board / sub-board
        return Y_BUS + LV_SUB_H * lv_level(items, bb)

    def no_toward(owner, other):      # owner's Notes: N.O. on the way to other
        return "no" in state_words(owner) and \
            has_word(owner.notes.lower(), (other.id.lower(),))

    def no_unnamed(owner, others):    # N.O. in Notes, no way named
        return "no" in state_words(owner) and not any(
            has_word(owner.notes.lower(), (o.id.lower(),)) for o in others)

    def y_rmu(r):                     # (top, bottom, bus) of an RMU box
        o = dy(r) + STEPUP_SHIFT
        return Y_RMU_TOP + o, Y_RMU_BOT + o, (Y_RMU_TOP + Y_RMU_BOT) / 2 + o

    def y_bus(b):                     # busbar level of an MV board
        return Y_MVBUS + dy(b) + STEPUP_SHIFT
    busbars = [items[i] for i in order if items[i].type == LV_BUSBAR]
    mvbs = [items[i] for i in order if items[i].type == MV_BUSBAR]
    rmus = [items[i] for i in order if items[i].type == RMU]
    lvsub = lv_subs(items, order)
    lvsub_mid = su_mid(items, order)
    txs = [items[i] for i in order
           if items[i].type == TRANSFORMER or i in lvsub]
    pumps = [items[i] for i in order if items[i].type == PUMP]
    mvs = [items[i] for i in order if items[i].type == MV_INCOMER]
    couplers = [items[i] for i in order if items[i].type == BUS_COUPLER]
    feeders = [items[i] for i in order
               if items[i].type in (FEEDER, MCC) + TERMINALS]

    site = info.get("site", "")
    title = f"{site} — Single Line Diagram (sketch)" if site \
        else "Single Line Diagram (sketch)"
    svg.layer = "frame"
    svg.text(24, DIAG_H - 26, title, size=16, anchor="start", bold=True)
    svg.layer = "drawing"

    # --- RMU-to-RMU link topology ----------------------------------------
    # straight cable between neighbouring boxes; when another RMU sits in
    # between (a ring closure), the cable loops over the top instead and
    # enters each box through an extra load-break-switch way
    side_links, ring_links = [], []
    ring_entries = {}  # rmu id -> extra top-entry x positions
    for rmu in rmus:
        for p in rmu.parents:
            if items[p].type != RMU or rmu.id in hang.get(p, ()):
                continue              # a hung RMU is fed from above
            if items[p].x is None or rmu.x is None:
                continue
            a, b = sorted((items[p], rmu), key=lambda r: r.x)
            tier = depth.get(a.id)
            between = any(r is not a and r is not b and a.x < r.x < b.x
                          and depth.get(r.id) == tier for r in rmus)
            if any(l[0] is a and l[1] is b for l in side_links) or \
                    any(l[3] == (a.id, b.id) for l in ring_links):
                continue              # the link written on both rows
            if between:
                xa, xb = a.x - 28, b.x + 28
                ring_entries.setdefault(a.id, []).append((xa, b.id))
                ring_entries.setdefault(b.id, []).append((xb, a.id))
                ring_links.append((xa, xb, Y_RMU_TOP + dy(a), (a.id, b.id)))
            else:
                side_links.append((a, b))

    # linked sides get a wider enclosure so the way switch fits inside
    pad_l = {r.id: 42 for r in rmus}
    pad_r = {r.id: 42 for r in rmus}
    for a, b in side_links:
        pad_r[a.id] = 64
        pad_l[b.id] = 64

    # --- RMU enclosures --------------------------------------------------
    rmu_box = {}
    for rmu in rmus:
        ways_in = [m for m in mvs if any(
            rmu.id == k.id for k in children_of(items, order, m.id))]
        ways_in += [items[t] for t in sus            # step-up columns
                    if any(rmu.id == k.id
                           for k in children_of(items, order, t))]
        ways_in += [tx for u, tx, l in links          # from a higher tier
                    if l.id == rmu.id and tx not in ways_in]
        ways_in += [items[g] for g in gens            # a generator of its own
                    if rmu.id in [b.id for b, _, _ in gens[g]]]
        boards_in = [items[p] for p in rmu.parents
                     if items[p].type == MV_BUSBAR]   # board-fed RMU
        # fed from the RMU above it (a spur or sub-ring off a ring)
        boards_in += [items[p] for p in rmu.parents
                      if rmu.id in hang.get(p, ())]
        ways_out = children_of(items, order, rmu.id,
                               {TRANSFORMER, PUMP} | set(LV_LOADS))
        ways_out += [k for k in children_of(items, order, rmu.id, {RMU})
                     if k.id in hang.get(rmu.id, ())]
        ways_out += [items[t] for t, (_, up) in su_mid(items, order).items()
                     if up.id == rmu.id and items[t].x is not None]
        def out_x(t):                 # a hung RMU leaves at its tee-off
            return rmu.tee.get(t.id, t.x)
        xs = [out_x(w) for w in ways_in + ways_out if w.x is not None] \
            or [rmu.x]
        xs = xs + [e[0] for e in ring_entries.get(rmu.id, [])]
        xs = xs + [rmu.land.get(b.id, rmu.x) for b in boards_in]
        left = min(xs) - pad_l[rmu.id]
        right = max(xs) + pad_r[rmu.id]
        bus_l, bus_r = min(xs) - 18, max(xs) + 18
        rmu_box[rmu.id] = (left, right, bus_l, bus_r)
        rt, rb, ymid = y_rmu(rmu)
        svg.rect(left, rt, right - left, rb - rt, sw=1.6, dash="7 5")
        # internal bus
        svg.line(bus_l, ymid, bus_r, ymid, w=3.4)
        # incoming ways from the top edge to the bus (default LBS,
        # overridden by the RMU row's Protection entry for that supply)
        def in_way(x, kind):
            if kind and kind != "lbs":
                svg.drop(x, rt, ymid, kind)
            else:
                svg.load_break_switch(x, rt + 12, ymid)
                svg.line(x, rt, x, rt + 12)
            svg.dot(x, ymid)
        for m in ways_in:
            in_way(m.x, prot_for(rmu, m.id)[1])
        for b in boards_in:          # the drop from above lands here
            in_way(rmu.land.get(b.id, rmu.x), prot_for(rmu, b.id)[1])
        # ring-closure entries come in through the top the same way
        for x_e, other in ring_entries.get(rmu.id, []):
            in_way(x_e, prot_for(rmu, other)[1]
                   if other in rmu.parents else None)
            if no_toward(rmu, items[other]):
                svg.text(x_e + 8, rt + 24, "N.O.", size=9, anchor="start")
        # outgoing ways from the bus to the bottom edge (default
        # fuse-switch, overridden by the fed item's Protection)
        for t in ways_out:
            kind = prot_for(t, rmu.id)[1]
            xo = out_x(t)
            if t.type == RMU and not kind:
                kind = "lbs"          # a cable to another RMU: a switch way
            if kind and kind != "fuse-switch":
                svg.drop(xo, ymid, rb, kind)
            else:
                svg.fuse_switch(xo, ymid + 4, rb - 8)
                svg.line(xo, ymid, xo, ymid + 4)
                svg.line(xo, rb - 8, xo, rb)
            svg.dot(xo, ymid)
        lbl = [rmu.id, rmu.desc,
               " ".join(v for v in (rmu.rating, rmu.voltage) if v)]
        # keep the label block clear of the next item on the same tier
        tier = depth.get(rmu.id)
        edges_r = [o.x_left if o.type == MV_BUSBAR else o.x - 60
                   for o in (items[q] for q in order)
                   if o.type in (RMU, MV_BUSBAR) and o is not rmu
                   and o.x is not None and depth.get(o.id) == tier
                   and o.x > rmu.x]
        edges_l = [o.x_right if o.type == MV_BUSBAR else o.x + 60
                   for o in (items[q] for q in order)
                   if o.type in (RMU, MV_BUSBAR) and o is not rmu
                   and o.x is not None and depth.get(o.id) == tier
                   and o.x < rmu.x]
        room_r = (min(edges_r) - right) if edges_r else 1e9
        room_l = (left - max(edges_l)) if edges_l else 1e9
        if room_r >= 130:
            lx, anc, ty = right + 10, "start", rt + 16
        elif room_l >= 130:
            lx, anc, ty = left - 10, "end", rt + 16
        else:                       # crowded tier: stack it above the box,
            # clear of the ring-closure lane at rt - 26
            lx, anc, ty = rmu.x, "middle", rt - 34 - 15 * (len(lbl) - 1)
        for i, t in enumerate(lbl):
            svg.text(lx, ty, t, anchor=anc, bold=(i == 0))
            ty += 15
        linked = [items[p] for p in rmu.parents if items[p].type == RMU] \
            + children_of(items, order, rmu.id, {RMU})
        if no_unnamed(rmu, linked):   # the open point is here, way unnamed
            svg.text(rmu.x + 12, rb + 14, "N.O.", size=9, anchor="start")

    # --- RMU-to-RMU interconnecting cables -------------------------------
    for a, b in side_links:  # a is the left box
        if a.id not in rmu_box or b.id not in rmu_box:
            continue
        y_link = y_rmu(a)[2]
        _, a_right, _, a_bus_r = rmu_box[a.id]
        b_left, _, b_bus_l, _ = rmu_box[b.id]
        svg.line(a_right, y_link, b_left, y_link, w=2)  # cable between boxes
        # the way switch inside each box, between the wall and the bus
        # (default LBS; the fed RMU's Protection entry can override it)
        for xe, xc, owner, other in ((a_right, a_bus_r, a, b),
                                     (b_left, b_bus_l, b, a)):
            kind = (prot_for(owner, other.id)[1]
                    if other.id in owner.parents else None) or "lbs"
            xm = (xe + xc) / 2
            gap = svg.device_h(kind, xm, y_link)
            lo, hi = min(xe, xc), max(xe, xc)
            svg.line(lo, y_link, xm - gap, y_link, w=2)
            svg.line(xm + gap, y_link, hi, y_link, w=2)
            if no_toward(owner, other):
                svg.text(xm, y_link + 22, "N.O.", size=9)
    for xa, xb, r_top, _ in ring_links:  # loop over the boxes in between
        y_ring = r_top - 26
        svg.line(xa, r_top, xa, y_ring)
        svg.line(xa, y_ring, xb, y_ring)
        svg.line(xb, y_ring, xb, r_top)
    # an RMU hung below its ring RMU: from the tee-off in the enclosure
    # down to a lane, across to the box below, and in through its top
    hang_runs = []
    for p, kids in hang.items():
        for c in sorted(kids, key=order.index):
            if items[p].x is None or items[c].x is None:
                continue
            x_t = items[p].tee.get(c, items[c].x)
            hang_runs.append(((p, c), x_t, items[c].x))
    hang_lane = {}
    for rb in {y_rmu(items[p])[1] for p in hang}:
        runs = [r for r in hang_runs
                if y_rmu(items[r[0][0]])[1] == rb and abs(r[1] - r[2]) >= 1]
        hang_lane.update(alloc_lanes(runs, [rb + 22, rb + 34, rb + 46]))
    for (p, c), x_t, x_c in hang_runs:
        rb, rt = y_rmu(items[p])[1], y_rmu(items[c])[0]
        if abs(x_t - x_c) < 1:
            svg.line(x_c, rb, x_c, rt)
        else:
            y_l = hang_lane[(p, c)]
            svg.line(x_t, rb, x_t, y_l)
            svg.line(x_t, y_l, x_c, y_l)
            svg.line(x_c, y_l, x_c, rt)

    # --- MV incomers -----------------------------------------------------
    for m in mvs:
        kids = children_of(items, order, m.id)
        # over a board on a lower tier the source tick sits just above it
        # rather than at the sheet top, so it crosses no other bar
        below = [dy(k) for k in kids if k.id in depth]
        y_top = Y_MV_TOP + (min(below) if below else 0)
        svg.line(m.x - 11, y_top, m.x + 11, y_top, w=3)  # source tick
        lbl = [m.id, m.desc, m.voltage]
        ty = y_top - Y_MV_TOP + Y_LABEL - 16
        for i, s in enumerate(lbl):
            svg.text(m.x, ty, s, size=11.5, bold=(i == 0))
            ty += 14
        for k in kids:
            if k.type == RMU:
                svg.line(m.x, y_top, m.x, y_rmu(k)[0])
            elif k.type == MV_BUSBAR:  # incoming device onto the board
                svg.drop(m.x, y_top, y_bus(k),
                         prot_for(k, m.id)[1] or "cb")
                svg.dot(m.x, y_bus(k))
            elif k.type == TRANSFORMER:  # direct feed, no RMU
                svg.line(m.x, y_top, m.x, Y_TX_C1 - TX_R)

    # --- MV switchboards -------------------------------------------------
    for mvb in mvbs:
        yb = y_bus(mvb)
        svg.line(mvb.x_left, yb, mvb.x_right, yb, w=5.5)
        raw, kind = prot_for(mvb)
        zone = raw if raw and kind is None else ""  # e.g. 87B differential
        lbl = " ".join(v for v in (mvb.id, mvb.desc, mvb.rating, mvb.voltage)
                       if v)
        if zone:
            lbl += " · " + zone
        svg.text(mvb.x_left, yb - 12, lbl, size=11.5, anchor="start",
                 bold=True)

    # --- feeds from an MV board down to a sub-board or an RMU -----------
    mv_feeds = []
    for oid in order:
        it = items[oid]
        if it.type not in (MV_BUSBAR, RMU) or it.x is None:
            continue
        for p in it.parents:
            par = items[p]
            if par.type != MV_BUSBAR:
                continue
            y_from = y_bus(par)
            y_to = y_bus(it) if it.type == MV_BUSBAR else y_rmu(it)[0]
            x_to = it.land.get(par.id, it.x)
            if it.type == MV_BUSBAR:
                x_to = min(max(x_to, it.x_left + 20), it.x_right - 20)
            x_from = min(max(x_to, par.x_left), par.x_right)
            mv_feeds.append((it, par, x_from, x_to, y_from, y_to))
    mv_lane = {}
    for band in {(f[4], f[5]) for f in mv_feeds}:
        runs = [((it.id, par.id), x_from, x_to)
                for it, par, x_from, x_to, y_from, y_to in mv_feeds
                if (y_from, y_to) == band and abs(x_from - x_to) >= 1]
        y_to = band[1]
        mv_lane.update(alloc_lanes(runs, [y_to - 30, y_to - 43, y_to - 56]))
    for it, par, x_from, x_to, y_from, y_to in mv_feeds:
        kind = ("cb" if it.type == RMU
                else prot_for(it, par.id)[1] or "cb")
        svg.dot(x_from, y_from)
        if abs(x_from - x_to) < 1:
            svg.drop(x_to, y_from, y_to, kind)
        else:                       # sub-board offset from the way
            y_mid = mv_lane[(it.id, par.id)]
            svg.drop(x_from, y_from, y_mid, kind)
            svg.line(x_from, y_mid, x_to, y_mid)
            svg.line(x_to, y_mid, x_to, y_to)
        if it.type == MV_BUSBAR:
            svg.dot(x_to, y_to)

    # --- pumps / motor loads --------------------------------------------
    for p in pumps:
        if p.x is None:
            continue
        # an MV motor sits in the transformer row; a motor fed from an LV
        # board or its own transformer hangs in the feeder band below it
        lv_par = [items[q] for q in p.parents
                  if items[q].type in (LV_BUSBAR, TRANSFORMER)]
        yc, r = ((Y_ARROW - 14, 14) if lv_par else (Y_PUMP, PUMP_R))
        if lv_par and lv_par[0].type == LV_BUSBAR:
            yc = lv_y(lv_par[0]) + 88 - 14
        vsd = "vsd" in state_words(p)
        for par in (items[q] for q in p.parents):
            if par.type == MV_BUSBAR:
                svg.drop(p.x, y_bus(par), Y_PUMP - PUMP_R,
                         prot_for(p, par.id)[1] or "cb")
                svg.dot(p.x, y_bus(par))
            elif par.type == RMU:   # the way device is inside the enclosure
                svg.line(p.x, y_rmu(par)[1], p.x, Y_PUMP - PUMP_R)
            elif par.type == LV_BUSBAR:          # LV motor off the board
                yb = lv_y(par)
                svg.dot(p.x, yb)
                svg.drop(p.x, yb, yc - r,
                         prot_for(p, par.id)[1] or "cb", ydev=yb + 30)
            elif par.type == TRANSFORMER:        # motor on its own transformer
                svg.drop(p.x, Y_TX_C2 + TX_R, yc - r,
                         prot_for(p, par.id)[1] or "cb")
        if vsd:                                  # drive box on the drop
            svg.vsd(p.x, yc - r - 14)
        svg.circle(p.x, yc, r, sw=2.2)
        svg.text(p.x, yc + 1, "M", size=13 if r > 13 else 11, bold=True)
        svg.text(p.x, yc + 13 if r > 13 else yc + 11, "3~", size=8.5)
        lbl = " · ".join(v for v in (p.id, p.desc, p.rating) if v)
        # an MV motor labels to the right when the next way leaves room;
        # otherwise (and always for LV motors) the label runs downward
        lines = [t for t in (p.id, p.desc, p.rating) if t]
        need = r + 10 + max((len(t) for t in lines), default=0) * 5.8 + 14
        room = 1e9
        if not lv_par:
            row = [o for o in order]
            nxt = [items[o].x for o in row
                   if items[o] is not p and items[o].x is not None
                   and items[o].x > p.x
                   and items[o].type in (TRANSFORMER, GENERATOR,
                                         PUMP, MV_BUSBAR)]
            nxt += [b.x_left for b in mvbs if b.x_left is not None
                    and b.x_left > p.x]
            room = (min(nxt) - p.x) if nxt else 1e9
        if not lv_par and room >= need:
            ty = yc - 6
            for t in lines:
                svg.text(p.x + r + 10, ty, t, size=11, anchor="start")
                ty += 14
        else:
            svg.text(p.x + 4, yc + r + 14, lbl, size=11, anchor="start",
                     rotate=90)

    # --- LV supply routes ------------------------------------------------
    # A board fed from several transformers gets one landing point per
    # supply, and every sideways run its own horizontal lane, so
    # cross-feeds between different MV boards never sit on one line.
    ytop_tx = Y_TX_C2 + TX_R
    routes, elbows = {}, []
    for tx in txs:
        if tx.x is None:
            continue
        for bb in children_of(items, order, tx.id, {LV_BUSBAR}):
            sup = [items[p] for p in bb.parents
                   if items[p].type == TRANSFORMER and items[p].x is not None]
            if len(sup) < 2:
                continue                      # single supply: handled below
            sup.sort(key=lambda t: t.x)
            i = [t.id for t in sup].index(tx.id)
            w = bb.x_right - bb.x_left
            x_land = bb.x_left + w * (i + 0.5) / len(sup)
            routes[(tx.id, bb.id)] = [x_land, None]
            if abs(x_land - tx.x) > 1:
                elbows.append(((tx.id, bb.id), tx.x, x_land))
    # every other run in the band between the transformer row and the LV
    # bar joins the same allocation: LV-fed step-ups taking supply from a
    # board, and transformers fed from a board with no output entered
    su_land = {}
    for tx_id, (src, up) in lvsub_mid.items():
        tx = items[tx_id]
        if tx.x is None or src.x_left is None:
            continue
        x_land = min(max(tx.x, src.x_left + 25), src.x_right - 25)
        su_land[tx_id] = x_land
        if abs(x_land - tx.x) >= 1:
            elbows.append((("su", tx_id), tx.x, x_land))
    high_runs = []
    for tx in txs:
        if (tx.x is None or tx.id in sus or tx.id in lvsub_mid
                or tx.id in ltx or gen_below(items, order, tx)):
            continue
        fed_lv = children_of(items, order, tx.id, {LV_BUSBAR})
        for p in tx.parents:
            par = items[p]
            if par.type != LV_BUSBAR or par.x_left is None:
                continue
            x_land = (par.x_right - 25 if tx.x > par.x else par.x_left + 25)
            if fed_lv:
                high_runs.append((("lvlv", tx.id, par.id), x_land, tx.x))
            else:
                elbows.append((("lvsrc", tx.id, par.id), tx.x, x_land))
    low_lane = alloc_lanes(elbows, [ytop_tx + 14, ytop_tx + 23, ytop_tx + 32])
    for key, y in low_lane.items():
        if key in routes:
            routes[key][1] = y
    c_top = Y_TX_C1 - TX_R
    high_lane = alloc_lanes(high_runs, [c_top - 22, c_top - 34,
                                        c_top - 88, c_top - 100])

    # --- generation sources feeding an LV board directly -----------------
    for g in (items[i] for i in order if items[i].type == GENERATOR):
        su_src = [k.id for i in order if items[i].type == TRANSFORMER
                  for k in children_of(items, order, i, {GENERATOR})]
        if g.x is None or g.id in [s.id for s in sus.values() if s] \
                or g.id in su_src or g.id in gens:
            continue                  # step-up sources are drawn in-column
        fed = [(b, k, c) for b, k, c in gen_feeds(items, order, g)
               if b.type == LV_BUSBAR and b.x_left is not None]
        # the circle stands over the (first) board it supplies
        yg = (lv_y(fed[0][0]) - (Y_BUS - Y_TX_C1)) if fed else Y_TX_C1
        svg.circle(g.x, yg + 13, 20, sw=2.2)
        svg.text(g.x, yg + 17, "G", size=13, bold=True)
        lbl = [g.id, g.desc, " ".join(v for v in (g.rating, g.voltage) if v)]
        ty = yg - 6
        for t in [t for t in lbl if t]:
            svg.text(g.x + 30, ty, t, anchor="start")
            ty += 15
        for bb, kind, cpl in fed:
            yb = lv_y(bb)
            svg.drop(g.x, yg + 33, yb, kind)
            svg.dot(g.x, yb)
            if cpl is not None:       # the changeover's own row, by its device
                ym = (yg + 33 + yb) / 2
                raw, k = prot_for(cpl)
                extra = raw if raw and k is None else ""
                svg.text(g.x + 16, ym - 2,
                         " ".join(v for v in (cpl.id, cpl.rating, extra)
                                  if v), size=11, anchor="start")
                svg.text(g.x + 16, ym + 12, cpl.notes, size=10,
                         anchor="start")

    # --- generators standing over MV gear as a supply --------------------
    for gid, fed in gens.items():
        g = items[gid]
        if g.x is None:
            continue
        b0 = fed[0][0]
        y_to0 = y_bus(b0) if b0.type == MV_BUSBAR else y_rmu(b0)[0]
        # on the top tier the column has headroom of its own; lower down
        # it fits in the 200 px tier gap, under the bar above
        if depth.get(b0.id, 0) == 0:
            y_gen = Y_GEN
        else:
            y_gen = y_to0 - (112 if b0.type == MV_BUSBAR else 82)
        svg.circle(g.x, y_gen, 20, sw=2.2)
        svg.text(g.x, y_gen + 4, "G", size=13, bold=True)
        lbl = [t for t in (g.id, g.desc, " ".join(
            v for v in (g.rating, g.voltage) if v)) if t]
        if b0.type == MV_BUSBAR:      # stacked above the circle, centred
            ty = y_gen - 28 - 14 * (len(lbl) - 1)
            for i, t in enumerate(lbl):
                svg.text(g.x, ty, t, size=11, bold=(i == 0))
                ty += 14
        else:                         # over an RMU: beside it, under the bar
            ty = y_gen - 12
            for t in lbl:
                svg.text(g.x + 30, ty, t, size=11, anchor="start")
                ty += 14
        for b, kind, cpl in fed:
            if b.type == MV_BUSBAR:
                y_to = y_bus(b)
                svg.drop(g.x, y_gen + 20, y_to, kind)
                svg.dot(g.x, y_to)
                if cpl is not None:
                    ym = (y_gen + 20 + y_to) / 2
                    svg.text(g.x + 16, ym - 2,
                             " ".join(v for v in (cpl.id, cpl.rating) if v),
                             size=11, anchor="start")
                    svg.text(g.x + 16, ym + 12, cpl.notes, size=10,
                             anchor="start")
            else:                     # the RMU draws its incoming way
                svg.line(g.x, y_gen + 20, g.x, y_rmu(b)[0])

    # --- transformers between two MV tiers -------------------------------
    for u, tx, l in links:
        if tx.x is None:
            continue
        y_u = y_bus(u) if u.type == MV_BUSBAR else y_rmu(u)[1]
        y_l = y_bus(l) if l.type == MV_BUSBAR else y_rmu(l)[0]
        c1 = (y_u + y_l) // 2 - 13     # integer: both engines round alike
        lbl = [tx.id, tx.desc,
               " ".join(v for v in (tx.rating, tx.voltage) if v)]
        svg.transformer(tx.x, [t for t in lbl if t], y=c1)
        # upper side: the board's outgoing way, whichever row named it
        kind_u = ((prot_for(tx, u.id)[1] if u.id in tx.parents
                   else prot_for(u, tx.id)[1]) or "cb")
        if u.type == MV_BUSBAR:
            if abs(tx.x - l.x) < 1 or u.x_left <= tx.x <= u.x_right:
                svg.drop(tx.x, y_u, c1 - TX_R, kind_u)
                svg.dot(tx.x, y_u)
            else:                     # a second link: dog-leg off the bar
                x_from = min(max(tx.x, u.x_left + 20), u.x_right - 20)
                y_lane = y_u + 40
                svg.drop(x_from, y_u, y_lane, kind_u)
                svg.dot(x_from, y_u)
                svg.line(x_from, y_lane, tx.x, y_lane)
                svg.line(tx.x, y_lane, tx.x, c1 - TX_R)
        else:                         # the RMU drew its tee-off to rb
            svg.line(tx.x, y_u, tx.x, c1 - TX_R)
        # lower side: the fed board's incomer
        kind_l = ((prot_for(l, tx.id)[1] if tx.id in l.parents
                   else prot_for(tx, l.id)[1]) or "cb")
        if l.type == MV_BUSBAR:
            svg.drop(tx.x, c1 + 27 + TX_R, y_l, kind_l)
            svg.dot(tx.x, y_l)
        else:                         # the RMU draws its incoming way
            svg.line(tx.x, c1 + 27 + TX_R, tx.x, y_l)

    # --- step-up columns: source -> transformer -> MV busbar / RMU -------
    for tx_id, src in sus.items():
        tx = items[tx_id]
        if tx.x is None:
            continue
        # the column stands over the board it feeds: on a lower tier the
        # whole column moves down with it
        fed0 = children_of(items, order, tx_id, {MV_BUSBAR, RMU})
        off = min((dy(f) for f in fed0), default=0)
        y_gen, y_c1, y_c2 = Y_GEN + off, Y_SU_C1 + off, Y_SU_C2 + off
        # the source symbol at the top of the column
        y_src_bot = y_gen + 20
        if src is None:
            y_src_bot = y_gen
        elif src.type == GENERATOR:
            svg.circle(tx.x, y_gen, 20, sw=2.2)
            svg.text(tx.x, y_gen + 4, "G", size=13, bold=True)
            lbl = [src.id, src.desc,
                   " ".join(v for v in (src.rating, src.voltage) if v)]
            ty = y_gen - 26
            for t in [t for t in lbl if t]:
                svg.text(tx.x + 30, ty, t, size=11, anchor="start")
                ty += 14
        elif src.type == LV_BUSBAR:   # a generation board
            svg.line(tx.x - 60, y_gen, tx.x + 60, y_gen, w=5.5)
            svg.text(tx.x - 60, y_gen - 12,
                     " ".join(v for v in (src.id, src.desc, src.rating,
                                          src.voltage) if v),
                     size=11.5, anchor="start", bold=True)
            y_src_bot = y_gen
        else:                          # an incomer row used as the source
            svg.line(tx.x - 11, y_gen, tx.x + 11, y_gen, w=3)
            lbl = [src.id, src.desc, src.voltage]
            ty = y_gen - 40
            for i, t in enumerate([t for t in lbl if t]):
                svg.text(tx.x, ty, t, size=11.5, bold=(i == 0))
                ty += 14
            y_src_bot = y_gen
        svg.line(tx.x, y_src_bot, tx.x, y_c1 - TX_R)
        svg.circle(tx.x, y_c1, TX_R, sw=2.2)
        svg.circle(tx.x, y_c2, TX_R, sw=2.2)
        lbl = [tx.id, tx.desc,
               " ".join(v for v in (tx.rating, tx.voltage) if v)]
        ty = y_c1 - 6
        for t in [t for t in lbl if t]:
            svg.text(tx.x + TX_R + 10, ty, t, anchor="start")
            ty += 15
        for fed in children_of(items, order, tx_id, {MV_BUSBAR, RMU}):
            y_to = y_bus(fed) if fed.type == MV_BUSBAR else y_rmu(fed)[0]
            svg.drop(tx.x, y_c2 + TX_R, y_to,
                     prot_for(fed, tx_id)[1] or "cb")
            if fed.type == MV_BUSBAR:
                svg.dot(tx.x, y_to)

    # --- step-ups from a real LV board: board below, MV gear above -------
    for tx_id, (src, up) in su_mid(items, order).items():
        tx = items[tx_id]
        if tx.x is None:
            continue
        svg.circle(tx.x, Y_TX_C1, TX_R, sw=2.2)
        svg.circle(tx.x, Y_TX_C2, TX_R, sw=2.2)
        lbl = [tx.id, tx.desc,
               " ".join(v for v in (tx.rating, tx.voltage) if v)]
        ty = Y_TX_C1 - 6
        for t in [t for t in lbl if t]:
            svg.text(tx.x + TX_R + 10, ty, t, anchor="start")
            ty += 15
        # down to the LV board that supplies it, routed onto its bar
        kind = prot_for(tx, src.id)[1] or "cb"
        x_land = su_land.get(tx_id, tx.x)
        if abs(x_land - tx.x) < 1:
            svg.drop(tx.x, Y_TX_C2 + TX_R, Y_BUS, kind)
        else:                         # the device is the board's, by the bar
            y_lane = low_lane[("su", tx_id)]
            svg.line(tx.x, Y_TX_C2 + TX_R, tx.x, y_lane)
            svg.line(tx.x, y_lane, x_land, y_lane)
            svg.drop(x_land, y_lane, Y_BUS, kind)
        svg.dot(x_land, Y_BUS)
        # up to the MV board / RMU it feeds
        if up.type == MV_BUSBAR:
            svg.drop(tx.x, y_bus(up), Y_TX_C1 - TX_R,
                     prot_for(up, tx.id)[1] or "cb")
            svg.dot(tx.x, y_bus(up))
        else:
            svg.line(tx.x, y_rmu(up)[1], tx.x, Y_TX_C1 - TX_R)

    # --- reversed step-ups: board on top, source below -------------------
    su_below = [items[i] for i in order
                if gen_below(items, order, items[i])]
    for tx in su_below:
        if tx.x is None:
            continue
        fed = [items[p] for p in tx.parents
               if items[p].type in (MV_BUSBAR, RMU)]
        for f in fed:
            y_from = y_bus(f) if f.type == MV_BUSBAR else y_rmu(f)[1]
            svg.drop(tx.x, y_from, Y_TX_C1 - TX_R,
                     prot_for(tx, f.id)[1] or "cb")
            if f.type == MV_BUSBAR:
                svg.dot(tx.x, y_from)
        svg.circle(tx.x, Y_TX_C1, TX_R, sw=2.2)
        svg.circle(tx.x, Y_TX_C2, TX_R, sw=2.2)
        lbl = [tx.id, tx.desc,
               " ".join(v for v in (tx.rating, tx.voltage) if v)]
        ty = Y_TX_C1 - 6
        for t in [t for t in lbl if t]:
            svg.text(tx.x + TX_R + 10, ty, t, anchor="start")
            ty += 15
        # the generation source hangs below the transformer
        kids = children_of(items, order, tx.id,
                           {GENERATOR, LV_BUSBAR, MV_INCOMER})
        src = kids[0] if kids else None
        y_src = Y_BUS
        if src is None:
            continue
        if src.type == GENERATOR:
            svg.line(tx.x, Y_TX_C2 + TX_R, tx.x, y_src - 20)
            svg.circle(tx.x, y_src, 20, sw=2.2)
            svg.text(tx.x, y_src + 4, "G", size=13, bold=True)
            lbl = [src.id, src.desc,
                   " ".join(v for v in (src.rating, src.voltage) if v)]
            ty = y_src - 6
            for t in [t for t in lbl if t]:
                svg.text(tx.x + 30, ty, t, size=11, anchor="start")
                ty += 14
        elif src.type == LV_BUSBAR:
            svg.line(tx.x, Y_TX_C2 + TX_R, tx.x, y_src)
            svg.dot(tx.x, y_src)     # the board itself is drawn with the
        else:                        # other LV busbars, on its own row
            svg.line(tx.x, Y_TX_C2 + TX_R, tx.x, y_src)
            svg.line(tx.x - 11, y_src, tx.x + 11, y_src, w=3)
            lbl = [src.id, src.desc, src.voltage]
            ty = y_src + 18
            for i, t in enumerate([t for t in lbl if t]):
                svg.text(tx.x, ty, t, size=11.5, bold=(i == 0))
                ty += 14

    # --- transformers ----------------------------------------------------
    for tx in txs:
        if (tx.x is None or tx.id in sus or tx.id in lvsub_mid
                or tx.id in ltx or gen_below(items, order, tx)):
            continue                  # drawn as a column of its own
        fed = children_of(items, order, tx.id, {LV_BUSBAR})
        lbl = [tx.id, tx.desc,
               " ".join(v for v in (tx.rating, tx.voltage) if v)]
        # transformers sharing a board label away from each other
        side = ("left" if any(len(bb.parents) > 1 and tx.x < bb.x
                              for bb in fed) else "right")
        if any(items[i].type == GENERATOR and items[i].x is not None
               and 0 < items[i].x - tx.x < 120 for i in order):
            side = "left"             # a generator stands to its right
        svg.transformer(tx.x, [s for s in lbl if s], side)
        for p in tx.parents:
            par = items[p]
            if par.type == RMU:
                svg.line(tx.x, y_rmu(par)[1], tx.x, Y_TX_C1 - TX_R)
            elif par.type == MV_BUSBAR:  # feeder device off the board
                svg.drop(tx.x, y_bus(par), Y_TX_C1 - TX_R,
                         prot_for(tx, par.id)[1] or "cb")
                svg.dot(tx.x, y_bus(par))
            elif par.type == LV_BUSBAR and par.x_left is not None:
                # supply comes back up from an LV board on the same row
                kind = prot_for(tx, par.id)[1] or "cb"
                x_land = (par.x_right - 25 if tx.x > par.x
                          else par.x_left + 25)
                if fed:      # LV/LV: supply in at the top, output below
                    y_lane = high_lane[("lvlv", tx.id, par.id)]
                    svg.drop(x_land, y_lane, Y_BUS, kind)
                    svg.line(x_land, y_lane, tx.x, y_lane)
                    svg.line(tx.x, y_lane, tx.x, Y_TX_C1 - TX_R)
                else:        # source below: the device is the board's
                    y_lane = low_lane[("lvsrc", tx.id, par.id)]
                    svg.line(tx.x, Y_TX_C2 + TX_R, tx.x, y_lane)
                    svg.line(tx.x, y_lane, x_land, y_lane)
                    svg.drop(x_land, y_lane, Y_BUS, kind)
                svg.dot(x_land, Y_BUS)
        if not any(items[p].type in (RMU, MV_BUSBAR, LV_BUSBAR)
                   and items[p].x is not None for p in tx.parents):
            svg.open_end(tx.x, Y_TX_C1 - TX_R, Y_TX_C1 - TX_R - 36,
                         "supply not defined")
        if earth_below(items, tx):    # earthing transformer / NER
            svg.line(tx.x, Y_TX_C2 + TX_R, tx.x, Y_TX_C2 + TX_R + 10)
            svg.resistor(tx.x, Y_TX_C2 + TX_R + 10)
        elif not fed and not children_of(items, order, tx.id,
                                         {PUMP, GENERATOR}):
            # the open side is whichever one the supply did not take
            if any(items[p].type == LV_BUSBAR for p in tx.parents):
                svg.open_end(tx.x, Y_TX_C1 - TX_R, Y_TX_C1 - TX_R - 36,
                             "outgoing not defined")
            else:
                svg.open_end(tx.x, Y_TX_C2 + TX_R, Y_TX_C2 + TX_R + 36,
                             "outgoing not defined")
        ytop = ytop_tx
        dual = [bb for bb in fed if (tx.id, bb.id) in routes]
        rest = [bb for bb in fed if (tx.id, bb.id) not in routes]
        for bb in dual:
            # a board with several incomers: each supply drops at its own
            # landing point, sideways runs on their own lanes
            kind = prot_for(bb, tx.id)[1] or "cb"
            x_land, y_lane = routes[(tx.id, bb.id)]
            if y_lane is None:
                svg.drop(tx.x, ytop, Y_BUS, kind)
            else:                     # the device is the board's, by the bar
                svg.line(tx.x, ytop, tx.x, y_lane)
                svg.line(tx.x, y_lane, x_land, y_lane)
                svg.drop(x_land, y_lane, Y_BUS, kind)
            svg.dot(x_land, Y_BUS)
        if len(rest) == 1 and abs(rest[0].x - tx.x) < 1:
            # straight drop to the busbar through the LV incomer device
            svg.drop(tx.x, ytop, Y_BUS,
                     prot_for(rest[0], tx.id)[1] or "cb")
            svg.dot(tx.x, Y_BUS)
        elif rest:
            # one transformer feeding several panels: split, then one
            # incomer device per panel
            ysplit = ytop + 32
            svg.line(tx.x, ytop, tx.x, ysplit)
            xs = [bb.x for bb in rest] + [tx.x]
            svg.line(min(xs), ysplit, max(xs), ysplit)
            for bb in rest:
                svg.dot(bb.x, ysplit)
                svg.drop(bb.x, ysplit, Y_BUS,
                         prot_for(bb, tx.id)[1] or "cb")
                svg.dot(bb.x, Y_BUS)

    # --- busbars ---------------------------------------------------------
    for bb in busbars:
        yb = lv_y(bb)
        svg.line(bb.x_left, yb, bb.x_right, yb, w=5.5)
        raw, kind = prot_for(bb)
        zone = raw if raw and kind is None else ""  # e.g. 87B differential
        lbl = " ".join(v for v in (bb.id, bb.desc, bb.rating, bb.voltage) if v)
        if zone:
            lbl += " · " + zone
        svg.text(bb.x_left, yb - 12, lbl, size=11.5, anchor="start",
                 bold=True)

    # --- sub-boards: fed from a feeder or straight from the board above --
    def two_devices(x, y0, y1, k0, k1, dash=None):
        """A drop with the outgoing device by the upper bar and, when
        the fed board names one, its incoming device by the lower bar."""
        g0 = svg.device(k0, x, y0 + 30)
        svg.line(x, y0, x, y0 + 30 - g0, dash=dash)
        if k1:
            g1 = svg.device(k1, x, y1 - 30)
            svg.line(x, y0 + 30 + g0, x, y1 - 30 - g1, dash=dash)
            svg.line(x, y1 - 30 + g1, x, y1, dash=dash)
        else:
            svg.line(x, y0 + 30 + g0, x, y1, dash=dash)

    split_done = set()
    for bb in busbars:
        if not is_sub_board(items, bb) or bb.x is None:
            continue
        y_sub = lv_y(bb)
        for p in bb.parents:
            par = items[p]
            if par.x is None:
                continue
            if par.type == LV_BUSBAR:
                y_par = lv_y(par)
                svg.dot(bb.x, y_par)
                two_devices(bb.x, y_par, y_sub,
                            prot_for(bb, p)[1] or "cb", None)
                svg.dot(bb.x, y_sub)
            elif par.type == FEEDER:
                pb = next((items[q] for q in par.parents
                           if items[q].type == LV_BUSBAR), None)
                if pb is None:
                    continue
                y_par = lv_y(pb)
                dash = "5 4" if "spare" in state_words(par) else None
                k0 = prot_for(par, pb.id)[1] or "cb"
                k1 = prot_for(bb, p)[1]
                svg.dot(par.x, y_par)
                if abs(par.x - bb.x) < 1:
                    two_devices(par.x, y_par, y_sub, k0, k1, dash)
                else:                 # several boards under one feeder
                    y_split = y_sub - 60
                    if p not in split_done:
                        split_done.add(p)
                        g0 = svg.device(k0, par.x, y_par + 30)
                        svg.line(par.x, y_par, par.x, y_par + 30 - g0,
                                 dash=dash)
                        svg.line(par.x, y_par + 30 + g0, par.x, y_split,
                                 dash=dash)
                    svg.line(par.x, y_split, bb.x, y_split, dash=dash)
                    if k1:
                        g1 = svg.device(k1, bb.x, y_sub - 30)
                        svg.line(bb.x, y_split, bb.x, y_sub - 30 - g1)
                        svg.line(bb.x, y_sub - 30 + g1, bb.x, y_sub)
                    else:
                        svg.line(bb.x, y_split, bb.x, y_sub)
                svg.dot(bb.x, y_sub)

    # --- bus couplers / ties ---------------------------------------------
    seen_pairs = {}
    for bc in couplers:
        if any(items[p].type == GENERATOR for p in bc.parents):
            continue                  # a changeover: drawn with the generator
        ends = [items[p] for p in bc.parents
                if items[p].type in (LV_BUSBAR, MV_BUSBAR)]
        if len(ends) != 2 or ends[0].type != ends[1].type:
            hint = (" (RMUs are tied with interconnecting cables - put the "
                    "other RMU in Feeds From instead)"
                    if any(items[p].type == RMU for p in bc.parents) else "")
            print(f"warning: bus coupler '{bc.id}' should feed from exactly "
                  f"two busbars of the same kind - skipping{hint}",
                  file=sys.stderr)
            continue
        a, b = sorted(ends, key=lambda e: e.x)
        if a.x_left is None or b.x_left is None:
            continue
        pair = tuple(sorted((a.id, b.id)))
        seen_pairs[pair] = seen_pairs.get(pair, 0) + 1
        if seen_pairs[pair] > 1:
            print(f"warning: bus coupler '{bc.id}' duplicates an earlier "
                  f"coupler between '{a.id}' and '{b.id}'", file=sys.stderr)
        ya = y_bus(a) if a.type == MV_BUSBAR else lv_y(a)
        yb = y_bus(b) if b.type == MV_BUSBAR else lv_y(b)
        raw, kind = prot_for(bc)
        dev = kind or "cb"
        extra = raw if raw and kind is None else ""
        lbl = " ".join(v for v in (bc.id, bc.rating, extra) if v)

        if abs(ya - yb) > 1:
            # the two boards sit on different levels: route clear of both
            x_link = max(a.x_right, b.x_right) + 34
            svg.dot(a.x_right, ya)
            svg.line(a.x_right, ya, x_link, ya, w=2)
            svg.drop(x_link, min(ya, yb), max(ya, yb), dev)
            svg.line(x_link, yb, b.x_right, yb, w=2)
            svg.dot(b.x_right, yb)
            ym = (ya + yb) / 2
            svg.text(x_link + 10, ym, lbl, size=11, anchor="start")
            svg.text(x_link + 10, ym + 14, bc.notes, size=10, anchor="start")
            continue

        blocking = [o for o in busbars + mvbs
                    if o is not a and o is not b and o.x_left is not None
                    and a.x_right < o.x < b.x_left
                    and abs((y_bus(o) if o.type == MV_BUSBAR else lv_y(o))
                            - ya) < 1]
        if blocking:
            # another board lies between: run the tie above the bar row
            y_lane = ya - 30
            xm = (a.x_right + b.x_left) / 2
            svg.dot(a.x_right, ya)
            svg.line(a.x_right, ya, a.x_right, y_lane, w=2)
            gap = svg.device_h(dev, xm, y_lane)
            svg.line(a.x_right, y_lane, xm - gap, y_lane, w=2)
            svg.line(xm + gap, y_lane, b.x_left, y_lane, w=2)
            svg.line(b.x_left, y_lane, b.x_left, yb, w=2)
            svg.dot(b.x_left, yb)
            svg.text(xm, y_lane - 10, lbl, size=11)
            svg.text(xm, y_lane - 24, bc.notes, size=10)
            continue

        xm = (a.x_right + b.x_left) / 2
        gap = svg.device_h(dev, xm, ya)
        svg.line(a.x_right, ya, xm - gap, ya, w=2)
        svg.line(xm + gap, ya, b.x_left, ya, w=2)
        svg.text(xm, ya + 24, lbl, size=11)
        svg.text(xm, ya + 38, bc.notes, size=10)

    # --- feeders ---------------------------------------------------------
    for f in feeders:
        if f.x is None:
            continue
        par = items[f.parents[0]] if f.parents else None
        kind = prot_for(f, par.id if par else None)[1] or "cb"
        dash = "5 4" if "spare" in state_words(f) else None
        lbl = " · ".join(v for v in (f.id, f.desc, f.rating) if v)
        if par is not None and par.type in (MV_BUSBAR, RMU):
            # an outgoing way of an MV board: arrow in the transformer row
            y_tip = Y_PUMP
            if par.type == MV_BUSBAR:
                y0 = y_bus(par)
                svg.dot(f.x, y0)
                y_end = (y_tip - 10 if f.type == FEEDER
                         else y_tip - 26 if f.type == MCC else y_tip - 24)
                svg.drop(f.x, y0, y_end, kind, dash=dash)
            else:                     # the way device is in the enclosure
                y0 = y_rmu(par)[1]
                y_end = (y_tip - 10 if f.type == FEEDER
                         else y_tip - 26 if f.type == MCC else y_tip - 24)
                svg.line(f.x, y0, f.x, y_end, dash=dash)
            if f.type == FEEDER:
                svg.arrow_down(f.x, y_tip)
                svg.text(f.x + 4, y_tip + 14, lbl, size=11, anchor="start",
                         rotate=90)
            elif f.type == MCC:
                svg.rect(f.x - 14, y_tip - 26, 28, 26, sw=2)
                svg.text(f.x, y_tip - 8, "MCC", size=8)
                svg.text(f.x + 4, y_tip + 14, lbl, size=11, anchor="start",
                         rotate=90)
            else:
                h = svg.terminal(f.type, f.x, y_tip - 24)
                svg.text(f.x + 4, y_tip - 24 + h + 12, lbl, size=11,
                         anchor="start", rotate=90)
            continue
        yb = lv_y(par) if par is not None and par.type == LV_BUSBAR \
            else Y_BUS
        y_dev, y_arrow, y_lbl = yb + 30, yb + 88, yb + 106
        svg.dot(f.x, yb)
        if f.type == MCC:  # motor control centre: box instead of arrow
            svg.drop(f.x, yb, y_arrow - 26, kind, ydev=y_dev, dash=dash)
            svg.rect(f.x - 14, y_arrow - 26, 28, 26, sw=2)
            svg.text(f.x, y_arrow - 8, "MCC", size=8)
        elif f.type in TERMINALS:
            svg.drop(f.x, yb, y_arrow - 24, kind, ydev=y_dev, dash=dash)
            h = svg.terminal(f.type, f.x, y_arrow - 24)
            y_lbl = y_arrow - 24 + h + 12
        elif sub_boards_of(items, order, f):
            # the drop carries on to the sub-board below (drawn with it);
            # the label sits beside the way, clear of that board
            svg.text(f.x + 8, y_dev + 30, lbl, size=11, anchor="start")
            continue
        else:
            svg.drop(f.x, yb, y_arrow - 10, kind, ydev=y_dev, dash=dash)
            svg.arrow_down(f.x, y_arrow)
        svg.text(f.x + 4, y_lbl, lbl, size=11, anchor="start", rotate=90)

    # --- title block -----------------------------------------------------
    svg.layer = "frame"
    tb_w, tb_h = 288, 96
    tb_x, tb_y = width - tb_w - 24, DIAG_H - tb_h - 20
    svg.rect(tb_x, tb_y, tb_w, tb_h, sw=1.5)
    svg.line(tb_x, tb_y + 24, tb_x + tb_w, tb_y + 24, w=1.5)
    svg.text(tb_x + 10, tb_y + 17, "SLD SKETCH — SITE SURVEY", size=12,
             anchor="start", bold=True)
    rows = [("Site", info.get("site", "")),
            ("Date", info.get("date", "")),
            ("By", info.get("surveyed by", info.get("by", ""))),
            ("Notes", info.get("notes", ""))]
    ty = tb_y + 40
    for k, v in rows:
        svg.text(tb_x + 10, ty, f"{k}:", size=11, anchor="start", bold=True)
        svg.text(tb_x + 58, ty, v[:44], size=11, anchor="start")
        ty += 16

    used = {items[i].type for i in order}
    if any(earth_below(items, items[i]) for i in order):
        used.add(EARTHING)
    svg.layer = "legend"
    draw_legend(svg, used)
    svg.layer = "drawing"

    return svg.document(width, DIAG_H + LEGEND_H)


# ---------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(
        description="Sketch a single-line diagram from a site-survey workbook.")
    ap.add_argument("workbook", help="input .xlsx file")
    ap.add_argument("-o", "--output",
                    help="output .svg file (default: <workbook>.svg)")
    ap.add_argument("--dxf", nargs="?", const="", metavar="FILE",
                    help="also write a DXF (R12) of the sketch with the "
                         "equipment table under it (default: <workbook>.dxf)")
    args = ap.parse_args()

    out = args.output or (args.workbook.rsplit(".", 1)[0] + ".svg")
    info, items, order = read_workbook(args.workbook)
    width = layout(items, order)
    svg = render(info, items, order, width)
    with open(out, "w", encoding="utf-8") as fh:
        fh.write(svg)
    print(f"wrote {out}  ({len(items)} items)")
    if args.dxf is not None:
        import sld_dxf
        dxf_out = args.dxf or (args.workbook.rsplit(".", 1)[0] + ".dxf")
        n = sld_dxf.write_dxf(dxf_out, info, items, order, width)
        print(f"wrote {dxf_out}  ({n} entities)")


if __name__ == "__main__":
    main()
